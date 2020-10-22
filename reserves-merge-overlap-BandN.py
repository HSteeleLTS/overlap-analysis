#!/usr/bin/env python3
import sys

import requests
import json
import os
import sys
import time
import csv
import re
import datetime

from tkinter.filedialog import askopenfilename

import pandas as pd
import numpy as np

import xml.etree.cElementTree as et
import pymarc as pym
import io
# from django.utils.encoding import smart_str

import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font


from urllib3.util.retry import Retry
from requests.adapters import HTTPAdapter


def ave_loop(record):
    perm_bool = False
    match_type = ""
    for ave in record.get_fields('AVE'):
        if 'm' in ave:
            if ('Covid-19'.lower() not in ave['m'].lower()):
                perm_bool = True
                match_type = "non-Covid"
                break
    if perm_bool == False:

        match_type = "Covid"

    return(match_type)

def ebook_match(record, m, t, a, yr, xm, xt, xtd, xa, xyr, u, ebook_match_on_list_counter, ebook_match_on_list_counter_without_year, temporary_collections_portfolio_counter_on_course, temporary_collections_counter_on_course_near_match, ebook_for_physical_counter, different_year_ebook_for_physical_counter, temporary_collections_portfolio_counter, temporary_collections_portfolio_counter_near_match, y, course_df, ebooks_to_add, ebooks_to_add_different_year, ebooks_we_need, covid_e_books_df, covid_e_books_near_match_df):

    master_match_type = ""
    loop_match_type = ""
    success = False

    # if ((str(t) == str(xt) or t == xtd) and (xa == a or xa in a or a in xa) and str(yr) == str(xyr)):
    #     loop_match_type = ave_loop(record)
    #     if loop_match_type == "non-Covid":
    #         ebook_for_physical_counter += 1
    #         base_series = course_df.iloc[y]
    #         add_series = pd.Series({'Match MMS ID': xml_mms_id, 'Match Title': xml_title, 'Match Author': xml_author, 'Match Publication Year': xml_year, 'Match URL or Collection': url})
    #         series_to_add = base_series.append(add_series)
    #         ebooks_to_add = ebooks_to_add.append(series_to_add, ignore_index=True)
    #     elif loop_match_type == 'Covid':
    #         base_series = course_df.iloc[y]
    #         add_series = pd.Series({'Match MMS ID': xml_mms_id, 'Match Title': xml_title, 'Match Author': xml_author, 'Match Publication Year': xml_year, 'Match URL or Collection': url})
    #         series_to_add = base_series.append(add_series)
    #         #series_to_add = base_series.append({'Match MMS ID': xml_mms_id, 'Match Title': xml_title, 'Match Author': xml_author, 'Match Publication Year': xml_year, 'Match URL or Collection': url})
    #         covid_e_books_df = covid_e_books_df.append(series_to_add, ignore_index=True)
    #         #ebooks_to_add = ebooks_to_add.append({'MMS ID': xml_mms_id, 'Title': xml_title, 'Author': xml_author, 'Publication Year': xml_year, 'URL or Collection': xml_url + "|" + collection}, ignore_index=True)
    #         temporary_collections_portfolio_counter += 1
    #     success = True
    if ((str(t) == str(xt) or t == xtd) and (xa == a or xa in a or a in xa)):
        loop_match_type = ave_loop(record)
        if loop_match_type == "non-Covid":
            base_series = course_df.iloc[y]
            add_series = pd.Series({'Match MMS ID': xm, 'Match Title': xt, 'Match Author': xa, 'Match Publication Year': xyr, 'Match URL or Collection': u})
            series_to_add = base_series.append(add_series)
            #series_to_add = base_series.append( Author': xml_author, 'Match Publication Year': xml_year, 'Match URL or Collection': url})
            ebooks_to_add_different_year = ebooks_to_add_different_year.append(series_to_add, ignore_index=True)
            #ebooks_to_add_different_year = ebooks_to_add_different_year.append({'Course Code': course_code, 'Course Name': course_name, 'MMS ID': xml_mms_id, 'Title': xml_title, 'Author': xml_author, 'Publication Year': xml_year, 'URL or Collection': url}, ignore_index=True)
            different_year_ebook_for_physical_counter += 1
        elif loop_match_type == 'Covid':
            base_series = course_df.iloc[y]
            add_series = pd.Series({'Match MMS ID': xm, 'Match Title': xt, 'Match Author': xa, 'Match Publication Year': xyr, 'Match URL or Collection': u})
            series_to_add = base_series.append(add_series)
            temporary_collections_portfolio_counter_near_match += 1

            covid_e_books_near_match_df = covid_e_books_near_match_df.append(series_to_add, ignore_index=True)
        success = True
    return([success, [ebook_match_on_list_counter, ebook_match_on_list_counter_without_year, temporary_collections_portfolio_counter_on_course, temporary_collections_counter_on_course_near_match, ebook_for_physical_counter, different_year_ebook_for_physical_counter, temporary_collections_portfolio_counter, temporary_collections_portfolio_counter_near_match], [ebooks_to_add, ebooks_to_add_different_year, ebooks_we_need, covid_e_books_df, covid_e_books_near_match_df]])


oDir = "./Output"
if not os.path.isdir(oDir) or not os.path.exists(oDir):
    os.makedirs(oDir)
start = datetime.datetime.now()
input_filename = askopenfilename(title="Select the Books to Order from the Barnes and Noble Proccess")

bndf = pd.read_excel(input_filename, dtype={'ISBN': 'str'})
with pd.option_context('display.max_columns', None):
    print(bndf)
bn_df_column_list_original = list(bndf.columns)
#drop ebook
bndf_column_list = ['BNDF Index', 'Title (Normalized)', 'MMS Id', 'ISBN', 'Processing Department', 'Course Name', 'Course Code', 'Author', 'Author (contributor)', 'Publication Date']
courses_df = pd.DataFrame(columns=bndf_column_list)



bndf['BNDF Index'] = bndf.index.tolist()
print(bndf)


print(courses_df)
sru_url_prefix_beginning = "https://tufts.alma.exlibrisgroup.com/view/sru/01TUN_INST?version=1.2&operation=searchRetrieve&recordSchema=marcxml&query="
#
for g in range(0, len(bndf)):
    isbn_bn = bndf.iloc[g]['ISBN']
    isbn_bn = str(isbn_bn)
    title_bn = bndf.iloc[g]['Title']
    author_bn = bndf.iloc[g]['Author']
    processing_dept_bn = "Tisch Library"
    course_code_bn = bndf.iloc[g]['Course']

    author_bn = author_bn.lower()
    # bn_title = bn_a + bn_b
    title_bn_original = title_bn
    title_bn = title_bn.lower()
    title_bn_original = title_bn

    ebook_boolean = False
    # print("Got through metadata assignment from BN file")
    # if '** E BOOK **' in isbn_bn or title_bn[0:3] == "EBK":
    #     ebook_boolean = True
    #     print("E book title: " + title_bn + "\n")
    #     # input("Press Any key")
    #     title_bn = title_bn[4:]
    #
    #     print("E book parsed title: " + title_bn + "\n")
    #     input("press any key")
    # else:
    #     ebook_boolean = False

    title_bn_dash = title_bn
    title_bn = re.sub(r'[^a-zA-Z0-9 ]', ' ', title_bn)
    title_bn_dash = re.sub(r'[^a-zA-Z0-9 -_]', ' ', title_bn_dash)
    title_bn = re.sub(r'\s{2,}', ' ', title_bn)
    title_bn_dash = re.sub(r'\s{2,}', ' ', title_bn_dash)
    title_for_query_input = '\"' + re.sub(r'\s', '%20', title_bn_dash) + '\"'
    # try:
    #     record_result = requests.get(sru_url_prefix_beginning + 'alma.title=' + str(title_for_query_input))
    # except:
    #     continue


    # s = requests.Session()
    #
    # retries = Retry(total=5,
    #                 backoff_factor=0.1,
    #                 status_forcelist=[ 500, 502, 503, 504 ])
    #
    # s.mount('https://', HTTPAdapter(max_retries=retries))

    # record_result = s.get(sru_url_prefix_beginning + 'alma.title=' + str(title_for_query_input))
    # try:

    record_result = requests.get(sru_url_prefix_beginning + 'alma.isbn=' + str(isbn_bn), timeout=10)
    # print(record_result.content)

    tree1 = et.ElementTree(et.fromstring(record_result.content))
    root1 = tree1.getroot()
    initial_search_success = True
    for elem1 in root1.iter():
        # print(elem1)
        if re.match(r'.+numberOfRecords.*', elem1.tag):
            # print("tag number of records for ISBN search.   Iteration: " + str(g))
            # print("input sheet row: " + str(bndf.iloc[g]))
            if int(elem1.text) == 0:
                print("no results for ISBN search.   Iteration: " + str(g))
                record_result = requests.get(sru_url_prefix_beginning + 'alma.title=' + str(title_for_query_input), timeout=10)
                tree2 = et.ElementTree(et.fromstring(record_result.content))
                root2 = tree2.getroot()
                for elem2 in root2.iter():
                    if re.match(r'.+numberOfRecords.*', elem2.tag):
                        print("tag number of records for title search.   Iteration: " + str(g))
                        if int(elem2.text) == 0:
                            print("No result in catalog.   Iteration: " + str(g))
                            courses_df = courses_df.append({'BNDF Index': g, 'Title (Normalized)': title_bn_original, 'MMS Id': "Not in Alma", 'ISBN': isbn_bn,'Processing Department': "Tisch Library", 'Course Name': re.sub(r'(\d{4}-\d{5}).*', r'\1', course_code_bn), 'Course Code': re.sub(r'(\d{4}-\d{5}).*', r'\1', course_code_bn), 'Author': author_bn, 'Author Contributor': "Not in Alma", 'Publication Date': "Not in Alma"}, ignore_index=True)
                            g += 1
                            initial_search_success = False
                            break
            break

    # sys.exit()

        # if (b'record' not in record_result.content):
        #     record_result = requests.get(sru_url_prefix_beginning + 'alma.title=' + str(title_for_query_input), timeout=10)
        #     if b'record' not in record_result.content:
        #         print("No result in catalog")
        #         courses_df = courses_df.append({'BNDF Index': g, 'Title (Normalized)': title_bn_original, 'MMS Id': "Not in Alma", 'ISBN': isbn_bn,'Processing Department': "Tisch Library", 'Course Name': re.sub(r'(\d{4}-\d{5}).*', r'\1', course_code_bn), 'Course Code': re.sub(r'(\d{4}-\d{5}).*', r'\1', course_code_bn), 'Author': author_bn, 'Author Contributor': "Not in Alma", 'Publication Date': "Not in Alma"}, ignore_index=True)
        #         g += 1
        #         continue



    # except:
    #     print("Couldn't get input URL")
    #     courses_df = courses_df.append({'BNDF Index': g, 'Title (Normalized)': title_bn_original, 'MMS Id': "Failed SRU Lookup", 'ISBN': isbn_bn,'Processing Department': "Tisch Library", 'Course Name': re.sub(r'(\d{4}-\d{5}).*', r'\1', course_code_bn), 'Course Code': re.sub(r'(\d{4}-\d{5}).*', r'\1', course_code_bn), 'Author': author_bn, 'Author Contributor': "Failed SRU Lookup", 'Publication Date': "Failed SRU Lookup"}, ignore_index=True)
    #     g += 1
    #     continue

    print("Input row: " + str(g))
    tree = et.ElementTree(et.fromstring(record_result.content))
    root = tree.getroot()
    for elem in root.iter():

        if re.match(r'.*record$', elem.tag):
            # print(elem.tag)
            bib_record = pym.parse_xml_to_array(io.StringIO(et.tostring(elem).decode('utf-8')))


            if '590' in bib_record[0]:
                if 'a' in bib_record[0]['590']:
                    if bib_record[0]['590']['a'].lower() == 'on the fly':
                        continue
            if '001' in bib_record[0]:

                bn_mms_id = ""


                bn_mms_id = bib_record[0]['001'].value()

                bn_a = ""
                bn_b = ""
                if 'a' in bib_record[0]['245']:
                    bn_a = bib_record[0]['245']['a']
                if 'b' in bib_record[0]['245']:
                    bn_b = bib_record[0]['245']['b']

                bn_title = bn_a + bn_b
                bn_title = bn_title.lower()
                bn_title = re.sub(r'\s\/$', '', bn_title)
                bn_title = re.sub(r'\'', ' ', bn_title)

                bn_title = re.sub(r'\.$', '', bn_title)
                bn_title = re.sub(r'^(the\s|a\s)', '', bn_title)
                bn_title = re.sub(r'\s{2,}', ' ', bn_title)
                bn_title_dash = bn_title
                bn_title = re.sub(r'[^a-zA-Z0-9 ]', ' ', bn_title)
                bn_title_dash = re.sub(r'[^a-zA-Z0-9 -_]', ' ', bn_title_dash)
                bn_title = re.sub(r'\s{2,}', ' ', bn_title)
                bn_title_dash = re.sub(r'\s{2,}', ' ', bn_title_dash)

                bn_author = ""
                bn_author_contributor = ""
                if '100' in bib_record[0]:
                    if 'a' in bib_record[0]['100']:
                        bn_author = bib_record[0]['100']['a']
                    if 'd' in bib_record[0]['100']:
                        bn_author += " " + bib_record[0]['100']['d']
                    if 'e' in bib_record[0]['100']:
                        bn_author += " " + bib_record[0]['100']['e']
                elif '110' in bib_record[0]:
                    if 'a' in bib_record[0]['110']:
                        bn_author = bib_record[0]['110']['a']
                    if 'd' in bib_record[0]['110']:
                        bn_author += " " + bib_record[0]['110']['d']
                    if 'e' in bib_record[0]['110']:
                        bn_author += " " + bib_record[0]['110']['e']
                elif '111' in bib_record[0]:
                    if 'a' in bib_record[0]['111']:
                        bn_author = bib_record[0]['111']['a']
                    if 'd' in bib_record[0]['111']:
                        bn_author += " " + bib_record[0]['111']['d']
                    if 'e' in bib_record[0]['111']:
                        bn_author += " " + bib_record[0]['111']['e']
                elif '700' in bib_record[0]:
                    if 'a' in bib_record[0]['700']:
                        bn_author_contributor = bib_record[0]['700']['a']
                    if 'd' in bib_record[0]['700']:
                        bn_author_contributor += " " + bib_record[0]['700']['d']
                    if 'e' in bib_record[0]['700']:
                        bn_author_contributor += " " + bib_record[0]['700']['e']
                elif '710' in bib_record[0]:
                    if 'a' in bib_record[0]['710']:
                        bn_author_contributor = bib_record[0]['710']['a']
                    if 'd' in bib_record[0]['710']:
                        bn_author_contributor += " " + bib_record[0]['710']['d']
                    if 'e' in bib_record[0]['710']:
                        bn_author_contributor += " " + bib_record[0]['710']['e']
                elif '711' in bib_record[0]:
                    if 'a' in bib_record[0]['711']:
                        bn_author_contributor = bib_record[0]['711']['a']
                    if 'd' in bib_record[0]['711']:
                        bn_author_contributor += " " + bib_record[0]['711']['d']
                    if 'e' in bib_record[0]['711']:
                        bn_author_contributor += " " + bib_record[0]['711']['e']


                xml_year = ""

                if '260' in bib_record[0]:
                    if 'c' in bib_record[0]['260']:
                        bn_year = bib_record[0]['260']['c']
                        bn_year = re.sub(r'\D', '', xml_year)
                elif '264' in bib_record[0]:
                    if 'c' in bib_record[0]['264']:
                        bn_year = bib_record[0]['264']['c']
                        bn_year = re.sub(r'\D', '', xml_year)

                # if (title_bn in bn_title or title_bn in bn_title_dash) and (author_bn in bn_author or author_bn in bn_author_contributor):

                courses_df = courses_df.append({'BNDF Index': g, 'Title (Normalized)': bn_title_dash, 'MMS Id': bn_mms_id, 'ISBN': isbn_bn,'Processing Department': "Tisch Library", 'Course Name': re.sub(r'(\d{4}-\d{5}).*', r'\1', course_code_bn), 'Course Code': re.sub(r'(\d{4}-\d{5}).*', r'\1', course_code_bn), 'Author': bn_author, 'Author Contributor': bn_author_contributor, 'Publication Date': bn_year}, ignore_index=True)
                # print("Match\n\n")

    g += 1
                    # continue
                # else:
                #     print("No match\n\n")
                #     courses_df = courses_df.append({'Title (Normalized)': title_bn_original, 'MMS Id': "Not in Alma", 'ISBN': isbn_bn,'Processing Department': "Tisch Library", 'Course Name': re.sub(r'(\d{4}-\d{5}).*', r'\1', course_code_bn), 'Course Code': re.sub(r'(\d{4}-\d{5}).*', r'\1', course_code_bn), 'Author': author_bn, 'Author Contributor': "Not in Alma", 'Publication Date': "Not in Alma"}, ignore_index=True)
                #     g += 1
                #     break


# z = 0

d = 0



courses_df = courses_df.drop_duplicates(['BNDF Index'])
courses_df = courses_df.reset_index()
with pd.option_context('display.max_columns', None):
    print(courses_df)

# sys.exit()
date = datetime.datetime.now().strftime("%m-%d-%Y")
filename = oDir + '/Electronic Copies from Barnes and Noble List - Physical Not in Alma - ' + date + '.xlsx'


wb = openpyxl.Workbook()
wb.save(filename)
workbook = load_workbook(filename)
writer = pd.ExcelWriter(filename, engine='openpyxl')
writer.book = workbook

course_df_list = []
# reserves_df_selected = reserves_df_selected[reserves_df_selected['Processing Department'] == 'Hirsh Health Sciences Reserves']
for course, course_df_1 in courses_df.groupby('Course Code'):
    course_df_list.append(course_df_1)
    # print(course_df_1)
## testing


x = 0
col_list = courses_df.columns.tolist()
output_cols = courses_df.columns.tolist()

output_cols.extend(['Match MMS ID', 'Match Title', 'Match Author', 'Match Publication Year', 'Match URL or Collection'])
counts_df = pd.DataFrame(columns=['Processing Department', 'Books on Course', 'Physical Books on Course', 'No Electronic Version for Physical Book', 'Electronic - Already on Course', 'Electronic - Already on Course - Different Year', 'Electronic - Already on Course - COVID Temporary Electronic Collection', 'Electronic - Already on Course - COVID Temporary Electronic Collection - Different Year', 'Electronic - In Collection - Add to Course', 'Electronic - In Collection - Potentially Add to Course - Different Year', 'Electronic - Temporarily in Collection', 'Electronic - Temporarily in Collection - Different Year'])
ebooks_to_add = pd.DataFrame(columns=output_cols)
ebooks_to_add_different_year = pd.DataFrame(columns=output_cols)

ebooks_we_need = pd.DataFrame(columns=bndf_column_list)


sru_url_prefix = "https://tufts.alma.exlibrisgroup.com/view/sru/01TUN_INST?version=1.2&operation=searchRetrieve&recordSchema=marcxml&query="

covid_e_books_df = pd.DataFrame(columns=output_cols)

covid_e_books_near_match_df = pd.DataFrame(columns=output_cols)
while x < len(course_df_list):

    y = 0
    temporary_collections_portfolio_counter = 0
    temporary_collections_portfolio_counter_on_course = 0
    temporary_collections_counter_on_course_near_match = 0
    temporary_collections_portfolio_counter_near_match = 0
    ebook_counter = 0
    ebook_for_physical_counter = 0
    different_year_ebook_for_physical_counter = 0
    ebook_match_on_list_counter_without_year = 0

    ebook_match_on_list_counter = 0
    ebooke_match_on_list_counter_without_year = 0
    no_match_counter = 0
    number_of_books_on_course = len(course_df_list[x])


    date = datetime.datetime.now().strftime("%m/%d/%Y")


    counts_per_course_list = []

    # print("Items on course: " + str(len(course_df_list[x])))
    course_ebook_count = 0
    course_code = course_df_list[x].iloc[0]['Course Code']
    course_name = course_df_list[x].iloc[0]['Course Name']

    print(course_code + "\n" + course_name + "\n")


    # physical_list = course_df_list[x][course_df_list[x]['E Book'] == False]
    #
    #
    # electronic_list = course_df_list[x][course_df_list[x]['E Book'] == True]

    number_of_physical_books_on_course = len(course_df_list[x])
    y = 0
    # print(physical_list)
    if len(course_df_list[x]) > 0:

        while y < len(course_df_list[x]):
            match = False
            course_df = course_df_list[x].copy()

            if course_df.iloc[y]['MMS Id'] == 'Not in Alma' or course_df.iloc[y]['MMS Id'] == 'Failed SRU Lookup':
                print("Not in Alma or Failed Lookup")
                # base_series = course_df.iloc[y]
                #add_series = pd.Series({'Course Code': course_code, 'Course Name': course_name, 'MMS ID': mms_id, 'Title': title, 'Author': author, 'Publication Year': year, 'Match MMS ID': xml_mms_id, 'Match Title': xml_title, 'Match Author': xml_author, 'Match Publication Year': xml_year, 'Match URL or Collection': url})
                # series_to_add = base_series.append(add_series)
                # series_to_add = base_series.append({'Match MMS ID': xml_mms_id, 'Match Title': xml_title, 'Match Author': xml_author, 'Match Publication Year': xml_year, 'Match URL or Collection': url})
                ebooks_we_need = ebooks_we_need.append(course_df.iloc[y], ignore_index=True)
                #ebooks_we_need = ebooks_we_need.append(course_df.iloc[y], ignore_index=True)
                no_match_counter += 1
                y += 1
                continue

            title = course_df.iloc[y]['Title (Normalized)']
            title_for_query = '\"' + re.sub(r'\s', '%20', title) + '\"'
            author = ""
            # print(str(author) + "Type: " + str(type(author)))
            if course_df.iloc[y]['Author'] != "" and course_df.iloc[y]['Author'] is not np.nan:
                #author = re.sub(r'([^,;]+,\s+[^,;],*\s*[^;,]*).+', r'\1', course_df.iloc[y]['Author'])
                author = re.sub(r'(,\sauthor\.?|,\scontributor\.?|\scontributor\.?|\sauthor\.?)', '', course_df.iloc[y]['Author'])
                author = author.lower()
            elif course_df.iloc[y]['Author (contributor)'] != "" and course_df.iloc[y]['Author (contributor)'] is not np.nan:
                author = re.sub(r'(,\sauthor|,\scontributor|\scontributor|\sauthor)', '', course_df.iloc[y]['Author (contributor)'])
                author = author.lower()
            #author_for_query = '\"' + re.sub(r'\s', '%20', author) + '\"'
            year = course_df.iloc[y]['Publication Date']
            year = re.sub(r'\D', '', year)
            mms_id = course_df.iloc[y]['MMS Id']



            # s2 = requests.Session()
            #
            # retries2 = Retry(total=5,
            #                 backoff_factor=0.1,
            #                 status_forcelist=[ 500, 502, 503, 504 ])
            #
            # s2.mount('https://', HTTPAdapter(max_retries=retries2))
            #
            # record_result_2 = s2.get(sru_url_prefix_beginning + 'alma.title=' + str(title_for_query_input), timeout=10)

            # try:
            #     tree3 = et.ElementTree(et.fromstring(record_result_2.content))
            #     root3 = tree3.getroot()
            #     for elem3 in root3.iter():
            #         if re.match(r'.numberOfRecords', elem3.tag):
            #             if int(elem3.text) == 0:
            #             print("No Match")
            #             ebooks_we_need = ebooks_we_need.append(course_df.iloc[y], ignore_index=True)
            #             y += 1
            #             continue
            record_result_2 = requests.get(sru_url_prefix_beginning + 'alma.title=' + str(title_for_query), timeout=10)
            # except:
            #     print("Couldn't get match URL")
            #     course_df.iloc[y]['MMS Id'] = course_df.iloc[y]['MMS Id'] + " not found in match search"
            #     ebooks_we_need = ebooks_we_need.append(course_df.iloc[y], ignore_index=True)
            #     y += 1
            #     continue
            #print(record_result.content)
            tree = et.ElementTree(et.fromstring(record_result_2.content))
            root = tree.getroot()
            # print(tree)
            # print(root)
            z = 0
            # result_df = pd.DataFrame(columns=['MMS Id', 'Title', 'Author', 'Year'])

            for elem in root.iter():


                if re.match(r'.*record$', elem.tag):
                    # print(elem.tag)
                    # print(et.tostring(elem))
                    bib_record = pym.parse_xml_to_array(io.StringIO(et.tostring(elem).decode('utf-8')))
                    # print(bib_record[0])
                    # test_file.write(str(bib_record[0]) + "\n")
                    # print(course_df.iloc[y])
                    # print(bib_record[0])
                    if '590' in bib_record[0]:
                        if 'a' in bib_record[0]['590']:
                            if bib_record[0]['590']['a'].lower() == 'on the fly':
                                continue
                    if '001' in bib_record[0]:

                        xml_mms_id = ""



                        xml_mms_id = bib_record[0]['001'].value()
                        # print('Analytics ID: ' + str(mms_id))
                        # print('XML MMS ID:   ' + str(xml_mms_id))


                        xml_url = ""

                        if '856' in bib_record[0]:
                            if 'u' in bib_record[0]['856']:
                                xml_url = bib_record[0]['856']['u']
                        a = ""
                        b = ""
                        if 'a' in bib_record[0]['245']:
                            a = bib_record[0]['245']['a']
                        if 'b' in bib_record[0]['245']:
                            b = bib_record[0]['245']['b']

                        xml_title = a + b
                        xml_title = xml_title.lower()
                        xml_title = re.sub(r'\s\/$', '', xml_title)
                        xml_title = re.sub(r'\'', ' ', xml_title)

                        xml_title = re.sub(r'\.$', '', xml_title)
                        xml_title = re.sub(r'^(the\s|a\s)', '', xml_title)
                        xml_title = re.sub(r'\s{2,}', ' ', xml_title)
                        xml_title_dash = xml_title
                        xml_title = re.sub(r'[^a-zA-Z0-9 ]', ' ', xml_title)
                        xml_title_dash = re.sub(r'[^a-zA-Z0-9 -_]', ' ', xml_title)
                        xml_title = re.sub(r'\s{2,}', ' ', xml_title)
                        xml_title_dash = re.sub(r'\s{2,}', ' ', xml_title_dash)

                        xml_author = ""

                        if '100' in bib_record[0]:
                            if 'a' in bib_record[0]['100']:
                                xml_author = bib_record[0]['100']['a']
                            if 'd' in bib_record[0]['100']:
                                xml_author += " " + bib_record[0]['100']['d']
                            if 'e' in bib_record[0]['100']:
                                xml_author += " " + bib_record[0]['100']['e']
                        elif '110' in bib_record[0]:
                            if 'a' in bib_record[0]['110']:
                                xml_author = bib_record[0]['110']['a']
                            if 'd' in bib_record[0]['110']:
                                xml_author += " " + bib_record[0]['110']['d']
                            if 'e' in bib_record[0]['110']:
                                xml_author += " " + bib_record[0]['110']['e']
                        elif '111' in bib_record[0]:
                            if 'a' in bib_record[0]['111']:
                                xml_author = bib_record[0]['111']['a']
                            if 'd' in bib_record[0]['111']:
                                xml_author += " " + bib_record[0]['111']['d']
                            if 'e' in bib_record[0]['111']:
                                xml_author += " " + bib_record[0]['111']['e']
                        elif '700' in bib_record[0]:
                            if 'a' in bib_record[0]['700']:
                                xml_author = bib_record[0]['700']['a']
                            if 'd' in bib_record[0]['700']:
                                xml_author += " " + bib_record[0]['700']['d']
                            if 'e' in bib_record[0]['700']:
                                xml_author += " " + bib_record[0]['700']['e']
                        elif '710' in bib_record[0]:
                            if 'a' in bib_record[0]['710']:
                                xml_author = bib_record[0]['710']['a']
                            if 'd' in bib_record[0]['710']:
                                xml_author += " " + bib_record[0]['710']['d']
                            if 'e' in bib_record[0]['710']:
                                xml_author += " " + bib_record[0]['710']['e']
                        elif '711' in bib_record[0]:
                            if 'a' in bib_record[0]['711']:
                                xml_author = bib_record[0]['711']['a']
                            if 'd' in bib_record[0]['711']:
                                xml_author += " " + bib_record[0]['711']['d']
                            if 'e' in bib_record[0]['711']:
                                xml_author += " " + bib_record[0]['711']['e']

                        # xml_author = xml_author.lower()
                        # xml_author = re.sub(r',$', '', xml_author)
                        xml_author = re.sub(r'(,\sauthor\.?|,\scontributor\.?|\scontributor\.?|\sauthor\.?)', '', xml_author)
                        xml_author = xml_author.lower()

                        xml_year = ""

                        if '260' in bib_record[0]:
                            if 'c' in bib_record[0]['260']:
                                xml_year = bib_record[0]['260']['c']
                                xml_year = re.sub(r'\D', '', xml_year)
                        elif '264' in bib_record[0]:
                            if 'c' in bib_record[0]['264']:
                                xml_year = bib_record[0]['264']['c']
                                xml_year = re.sub(r'\D', '', xml_year)

                        collection = ""
                        type = ""
                        if 'AVE' in bib_record[0]:
                            if 'm' in bib_record[0]['AVE']:
                                collection = bib_record[0]['AVE']['m']
                                type = 'temp'
                        url = ""
                        if collection != "" and xml_url != "":
                            url =  xml_url + "|" + collection
                        elif collection != "":
                            url = collection
                        elif xml_url != "":
                            url = xml_url

                        electronic_h_bool = False
                        if 'h' in bib_record[0]['245']:
                            if 'electronic resource' in bib_record[0]['245']['h'] or 'Electronic resource' in bib_record[0]['245']['h']:
                                electronic_h_bool = True
                                type = 'electronic_245_h'


                        if '655' in bib_record[0] and "Electronic books".lower() in bib_record[0]['655']['a'].lower():
                            type = 'electronic_655'



                        if 'AVE' in bib_record[0]:
                            # print('AVE in bib record')
                            # print(bib_record[0])
                            return_list = ebook_match(bib_record[0], mms_id, title, author, year, xml_mms_id, xml_title, xml_title_dash, xml_author, xml_year, url, ebook_match_on_list_counter, ebook_match_on_list_counter_without_year, temporary_collections_portfolio_counter_on_course, temporary_collections_counter_on_course_near_match, ebook_for_physical_counter, different_year_ebook_for_physical_counter, temporary_collections_portfolio_counter, temporary_collections_portfolio_counter_near_match, y, course_df, ebooks_to_add, ebooks_to_add_different_year, ebooks_we_need, covid_e_books_df, covid_e_books_near_match_df)
                            match = return_list[0]
                            counts_return = return_list[1]
                            dataframe_return = return_list[2]

                            #counts
                            ebook_match_on_list_counter = counts_return[0]
                            ebook_match_on_list_counter_without_year = counts_return[1]
                            temporary_collections_portfolio_counter_on_course = counts_return[2]
                            temporary_collections_counter_on_course_near_match = counts_return[3]
                            ebook_for_physical_counter = counts_return[4]
                            different_year_ebook_for_physical_counter = counts_return[5]
                            temporary_collections_portfolio_counter = counts_return[6]
                            temporary_collections_portfolio_counter_near_match = counts_return[7]

                            #dataframes
                            ebooks_to_add = dataframe_return[0]
                            ebooks_to_add_different_year = dataframe_return[1]
                            ebooks_we_need = dataframe_return[2]
                            covid_e_books_df = dataframe_return[3]
                            covid_e_books_near_match_df = dataframe_return[4]

                            if match == True:
                                z += 1
                                break




                z += 1
            #     if (xml_mms_id != mms_id and type != ""):
            #         print("Match after of XML loop:" + str(match))
            # if (xml_mms_id != mms_id and type != ""):
            #     print("Match after else statement: " + str(match))
            #     print("XML:        " + str(xml_mms_id) + "|" + str(xml_title) + "|" + str(xml_author) + "|" + str(xml_year) + "|" + type)
            #     print("Analytics:  " + str(mms_id) + "|" + str(title) + "|" + str(author) + "|" + str(year))
            if not match:
                print("Defaulted into not found")
                # base_series = course_df.iloc[y]
                #add_series = pd.Series({'Course Code': course_code, 'Course Name': course_name, 'MMS ID': mms_id, 'Title': title, 'Author': author, 'Publication Year': year, 'Match MMS ID': xml_mms_id, 'Match Title': xml_title, 'Match Author': xml_author, 'Match Publication Year': xml_year, 'Match URL or Collection': url})
                # series_to_add = base_series.append(add_series)
                # series_to_add = base_series.append({'Match MMS ID': xml_mms_id, 'Match Title': xml_title, 'Match Author': xml_author, 'Match Publication Year': xml_year, 'Match URL or Collection': url})
                index_value = course_df.iloc[y]['BNDF Index']
                ebooks_we_need = ebooks_we_need.append(course_df.iloc[y], ignore_index=True)
                #ebooks_we_need = ebooks_we_need.append(course_df.iloc[y], ignore_index=True)
                no_match_counter += 1

            #print(ebooks_on_this_course)
            y += 1

    course_name = course_df_list[x].iloc[0]['Course Name']
    course_code = course_df_list[x].iloc[0]['Course Code']
    processing_dept = course_df_list[x].iloc[0]['Processing Department']


    #counts_file = open(oDir + '/Counts - ' + str(course_name) + ' - ' + date + '.txt', 'w+')

    counts_df = counts_df.append({'Processing Department': processing_dept, 'Course Name': course_name, 'Course Code': course_code, 'Books on Course': number_of_books_on_course,  'Physical Books on Course' :number_of_physical_books_on_course,  'No Electronic Version for Physical Book': no_match_counter, 'Electronic - Already on Course': ebook_match_on_list_counter, 'Electronic - Already on Course - Different Year': ebook_match_on_list_counter_without_year, 'Electronic - Already on Course - COVID Temporary Electronic Collection': temporary_collections_portfolio_counter_on_course,  'Electronic - Already on Course - COVID Temporary Electronic Collection - Different Year': temporary_collections_counter_on_course_near_match, 'Electronic - In Collection - Add to Course': ebook_for_physical_counter, 'Electronic - In Collection - Potentially Add to Course - Different Year': different_year_ebook_for_physical_counter, 'Electronic - Temporarily in Collection': temporary_collections_portfolio_counter, 'Electronic - Temporarily in Collection - Different Year': temporary_collections_portfolio_counter_near_match}, ignore_index=True)


    x += 1
# This section moves the last two columns, course data, to the first two positions, and puts proc dept in beginning
cols = counts_df.columns.tolist()
cols = cols[-1:] + cols[:-1]
counts_df = counts_df[cols]
cols = cols[-1:] + cols[:-1]
counts_df = counts_df[cols]
cols.remove('Processing Department')
cols.insert(0, 'Processing Department')
counts_df = counts_df[cols]


# ebooks_to_add = ebooks_to_add.drop_duplicates(subset=['MMS Id', 'Match MMS ID', 'Match URL or Collection'])
# ebooks_to_add_different_year = ebooks_to_add_different_year.drop_duplicates(subset=['MMS Id', 'Match MMS ID', 'Match URL or Collection'])
# covid_e_books_df = covid_e_books_df.drop_duplicates(subset=['MMS Id', 'Match MMS ID', 'Match URL or Collection'])
# covid_e_books_near_match_df = covid_e_books_near_match_df.drop_duplicates(subset=['MMS Id', 'Match MMS ID', 'Match URL or Collection'])

ebooks_to_add = ebooks_to_add.drop_duplicates(subset=['BNDF Index'])
ebooks_to_add_different_year = ebooks_to_add_different_year.drop_duplicates(subset=['BNDF Index'])
covid_e_books_df = covid_e_books_df.drop_duplicates(subset=['BNDF Index'])
covid_e_books_near_match_df = covid_e_books_near_match_df.drop_duplicates(subset=['BNDF Index'])


counts_df.to_excel(writer, sheet_name='Counts', index=False, engine='openpyxl')
ebooks_to_add.to_excel(writer, sheet_name='InRepo', index=False, engine='openpyxl')
ebooks_to_add_different_year.to_excel(writer, sheet_name='InRepoDifferentYear', index=False, engine='openpyxl')
covid_e_books_df.to_excel(writer, sheet_name='TempColl', index=False, engine='openpyxl')
covid_e_books_near_match_df.to_excel(writer, sheet_name='TempCollDiffYear', index=False, engine='openpyxl')
ebooks_we_need.to_excel(writer, sheet_name='NeededNotInTufts', index=False, engine='openpyxl')

ebooks_we_need_unique = ebooks_we_need.drop_duplicates(subset=['ISBN'])

# print("\n\n\n")
# print(ebooks_we_need_unique)

ebooks_we_need_unique.to_excel(writer, sheet_name='UniquePrchsList', index=False, engine='openpyxl')




for sheet in writer.sheets:
    e = 0
    for column in writer.sheets[sheet].iter_cols():
        writer.sheets[sheet].column_dimensions[get_column_letter(e + 1)].width = "20"
        e += 1
f = 0

for sheet in writer.sheets:
    writer.sheets[sheet].freeze_panes = 'A2'
    for row in writer.sheets[sheet]:
        # print(row)
        if f == 0:

            for cell in row:
                #cell.style.alignment.wrap_text=True
                cell.alignment = Alignment(wrap_text=True)
                cell.font = Font(bold=True)


        f += 1

# sums
counts_max_row = writer.book['Counts'].max_row

for x in range(1, 13):
    writer.sheets['Counts'].cell(row = counts_max_row + 2, column = x).font = Font(bold=True)

writer.sheets['Counts'].cell(row = counts_max_row + 2, column = 1).value = "Totals"
writer.sheets['Counts'].cell(row = counts_max_row + 2, column = 4).value = '= SUM(D1:D' + str(counts_max_row) + ')'
writer.sheets['Counts'].cell(row = counts_max_row + 2, column = 5).value = '= SUM(E1:E' + str(counts_max_row) + ')'
writer.sheets['Counts'].cell(row = counts_max_row + 2, column = 6).value = '= SUM(F1:F' + str(counts_max_row) + ')'
writer.sheets['Counts'].cell(row = counts_max_row + 2, column = 7).value = '= SUM(G1:G' + str(counts_max_row) + ')'
writer.sheets['Counts'].cell(row = counts_max_row + 2, column = 8).value = '= SUM(H1:H' + str(counts_max_row) + ')'
writer.sheets['Counts'].cell(row = counts_max_row + 2, column = 9).value = '= SUM(I1:I' + str(counts_max_row) + ')'
writer.sheets['Counts'].cell(row = counts_max_row + 2, column = 10).value = '= SUM(J1:J' + str(counts_max_row) + ')'
writer.sheets['Counts'].cell(row = counts_max_row + 2, column = 11).value = '= SUM(K1:K' + str(counts_max_row) + ')'
writer.sheets['Counts'].cell(row = counts_max_row + 2, column = 12).value =  '= SUM(L1:L' + str(counts_max_row) + ')'
writer.sheets['Counts'].cell(row = counts_max_row + 2, column = 11).value = '= SUM(M1:M' + str(counts_max_row) + ')'
writer.sheets['Counts'].cell(row = counts_max_row + 2, column = 12).value =  '= SUM(N1:N' + str(counts_max_row) + ')'

#
#
first_sheet = workbook.get_sheet_by_name('Sheet')



workbook.remove_sheet(first_sheet)
writer.save()
workbook.save(filename)

# print(proc_dept_df_list)
# print(proc_dept_df_list[d])

d += 1

end = datetime.datetime.now() - start
# test_file.close()

print("Execution time: " + str(end) + "\n")
