#!/usr/bin/env python3
# -*- coding: utf-8 -*-
######################################################################################################
######################################################################################################
######################################################################################################
######################################################################################################
########
########
########    Author:           Henry Steele, Library Technology Services, Tufts University
########    Name of Program:  Citations
########	Files:			  citations.py, Scripts/functions.py
########    Date created:     2018-12
########
########    Purpose:
########      - To create a series of word documents that contain bibliographies of all the Titles
########        purchased in a given fiscal year for a given library (Tisch or Ginn)
########
########    Method:
########      - provide library and fiscal year prompt
########      - program retrives the appropriate Analytics report:
########          - either/or
########              - /shared/Tufts University/Reports/Collections/Gift Funds/Titles Purchased with Gift Funds - Tisch - Generic for Script
########              - /shared/Tufts University/Reports/Collections/Gift Funds/Titles Purchased with Gift Funds - Ginn - Generic for Script
########          - outputs:
########              - MMS Id
########              - fund
########          - filters on
########              - "MMS Id is not equal to / is not in  -1"
########              - (Tisch) "AND Fund Ledger Code is equal to / is in  dalex; dalel; daron; dbarr; dcamp; dchri; dcros; dduke; dfitc; dgiff; dgonz; dgord; dhaly; dharo; dloeb; dmeas; dnewh; dpall; dprit; drose; drosg; dshap; dsper; dtisc; dwill; dfox; docon; dcohe; dargo; dblak; dmarc"
########              - OR (Ginn) "Fund Ledger Name is equal to / is in  Bradley - Books; Cabot - Books; Fares - Books; Hay - Books; Imlah - Books; Maney - Books; Raanan - Books; Salacuse - Books; Saskawa-NPP - Books"
########              - "AND Transaction Date is prompted"
########                  - this is passed as a 'saw' XML filter in the URL that encodes the date range
########      - retrieves the XML report, iterates through and parses MMS Id and fund
########      - performs an SRU search by MMS Id
########      - parses out relevant title, author, and pulication information field from bib XML
########          + MMS Id
########          + Main entry Author (MARC 100|a)
########          + Main entry Author relator (MARC 100|e)
########          + Second author (MARC 110|a)
########          + Second author relator (MARC 110|e)
########          + Corporate author (MARC 700|a)
########          + Corporate author relator (MARC 700|e)
########          + Second corporate author (MARC 710|a)
########          + Second corporate author relator (MARC 710|e)
########          + Title (MARC 245|a)
########          + Subtitle (MARC 245|b)
########          + Place of publication (MARC 260|a)
########          + Name of publisher (MARC 260|b)
########          + Date of publication (MARC 260|c)
########          + Place of second publication (MARC 264|a)
########          + Name of second publisher (MARC 264|b)
########          + Date of second publication (MARC 264|c)
########      - turns this data into a ".bib" BibTex file
########      - uses locally included citeproc.py module to create bibliography, and local docx module to write to Word
########      - These have to be locally included because I had to change some of the internals of these pacakges to handle UTF-8 encoding
########
########    Dependences:
########      - in "requirements.txt"
########         - tkinter.filedialog import askopenfilename
########         - from django.utils.encoding import smart_bytes
########         - import pandas as pd
########         - import numpy as np
########         - import docx
########         - import xml.etree.ElementTree as et
########         - various citeproc-py methods
########
########    Output:
########      - "Processing" directory contains intermediate ".bib" file, which is in BibTex that citeproc
########      - "Output" directory contains final Word .docx file
########
########    Troubleshooting:
########      - The most likely errors you will encounter will be with encoding.
########        The script translates everythign into UTF-8 so foreign characters shouldn't be a problem,
########        but if you do run into issues you may want to exempt the individual bib record from input files_to_ignore
########        (in /Processing), commend out the part of the code all the way up to where they are created, and rerun.
########        Or fix the records and wait a day for a new Analtics report
import sys

import requests
import json
import os
import sys
import time
import csv
import re
import datetime

from tkinter.filedialog import askopenfilename

import pandas as pd
import numpy as np

import xml.etree.cElementTree as et
import pymarc as pym
import io
# from django.utils.encoding import smart_str

import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font

from urllib.parse import urlencode, quote

#####################################################
#####################################################
########    function to determine if an electronic
########    is in a temporary Covid collection
def ave_loop(record):
    perm_bool = False
    match_type = ""
    for ave in record.get_fields('AVE'):
        if 'm' in ave:
            if ('Covid-19'.lower() not in ave['m'].lower()):
                perm_bool = True
                match_type = "non-Covid"
                break
    if perm_bool == False:

        match_type = "Covid"

    return(match_type)

#####################################################
#####################################################
########    Using an Alma SRU search, get the XML
########    results for a title search, and parse
########    with Python Element Tree and PyMARC
########    and parse out fields to compare
########    to incoming records
def get_xml(title_for_query, sru_url_prefix):
    # print("Title for query:" + str(title_for_query))
    record_result = requests.get(sru_url_prefix + 'alma.title=' + title_for_query)

    tree = et.ElementTree(et.fromstring(record_result.content))
    # print("record result content: \n" + str(record_result.content))
    root = tree.getroot()

    try:
        records[:] = []
    except:
        records = []
    # print("\n\n\nRecords at beginning.  Should be empty: " + str(records) + "\n\n\n")

    for elem in tree.iter():

        if re.match(r'.*record$', elem.tag):
            #     print(elem.tag)
            # print(et.tostring(elem))
            bib_record = pym.parse_xml_to_array(io.StringIO(et.tostring(elem).decode('utf-8')), strict=False)


            for record in bib_record:

                # print(record)
                if record is None:
                    continue

                if '590' in record:
                    if 'a' in record['590']:
                        if record['590']['a'].lower() == 'on the fly':
                            continue
                            y += 1
                if '001' in record:

                    xml_mms_id = ""



                    xml_mms_id = record['001'].value()



                    xml_url = ""

                    if '856' in record:
                        if 'u' in record['856']:
                            xml_url = record['856']['u']
                    a = ""
                    b = ""
                    if 'a' in record['245']:
                        a = record['245']['a']
                    if 'b' in record['245']:
                        b = record['245']['b']

                    xml_title = a + b
                    xml_title = xml_title.lower()
                    xml_title = re.sub(r'\s\/$', '', xml_title)
                    xml_title = re.sub(r'\'', ' ', xml_title)

                    xml_title = re.sub(r'\.$', '', xml_title)
                    xml_title = re.sub(r'^(the\s|a\s)', '', xml_title)
                    xml_title = re.sub(r'\s{2,}', ' ', xml_title)
                    xml_title_dash = xml_title
                    xml_title = re.sub(r'[^a-zA-Z0-9 ]', ' ', xml_title)
                    xml_title_dash = re.sub(r'[^a-zA-Z0-9 -_]', ' ', xml_title)
                    xml_title = re.sub(r'\s{2,}', ' ', xml_title)
                    xml_title_dash = re.sub(r'\s{2,}', ' ', xml_title_dash)

                    xml_author_100s = ""
                    xml_author_700s = ""

                    if '100' in record:
                        if 'a' in record['100']:
                            xml_author_100s = record['100']['a']
                        if 'd' in record['100']:
                            xml_author_100s += " " + record['100']['d']
                        if 'e' in record['100']:
                            xml_author_100s += " " + record['100']['e']
                    elif '110' in record:
                        if 'a' in record['110']:
                            xml_author_100s = record['110']['a']
                        if 'd' in record['110']:
                            xml_author_100s += " " + record['110']['d']
                        if 'e' in record['110']:
                            xml_author_100s += " " + record['110']['e']
                    elif '111' in record:
                        if 'a' in record['111']:
                            xml_author_100s = record['111']['a']
                        if 'd' in record['111']:
                            xml_author_100s += " " + record['111']['d']
                        if 'e' in record['111']:
                            xml_author_100s += " " + record['111']['e']
                    if '700' in record:
                        if 'a' in record['700']:
                            xml_author_700s = record['700']['a']
                        if 'd' in record['700']:
                            xml_author_700s += " " + record['700']['d']
                        if 'e' in record['700']:
                            xml_author_700s += " " + record['700']['e']
                    elif '710' in record:
                        if 'a' in record['710']:
                            xml_author_700s = record['710']['a']
                        if 'd' in record['710']:
                            xml_author_700s += " " + record['710']['d']
                        if 'e' in record['710']:
                            xml_author_700s += " " + record['710']['e']
                    elif '711' in record:
                        if 'a' in record['711']:
                            xml_author_700s = record['711']['a']
                        if 'd' in record['711']:
                            xml_author_700s += " " + record['711']['d']
                        if 'e' in record['711']:
                            xml_author_700s += " " + record['711']['e']

                    # xml_author = xml_author.lower()
                    # xml_author = re.sub(r',$', '', xml_author)
                    xml_author_100s = re.sub(r'(,\sauthor\.?|,\scontributor\.?|\scontributor\.?|\sauthor\.?)', '', xml_author_100s)
                    xml_author_100s = xml_author_100s.lower()

                    xml_author_700s = re.sub(r'(,\sauthor\.?|,\scontributor\.?|\scontributor\.?|\sauthor\.?)', '', xml_author_700s)
                    xml_author_700s = xml_author_700s.lower()
                    xml_year = ""

                    if '260' in record:
                        if 'c' in record['260']:
                            xml_year = record['260']['c']
                            xml_year = re.sub(r'\D', '', xml_year)
                    elif '264' in record:
                        if 'c' in record['264']:
                            xml_year = record['264']['c']
                            xml_year = re.sub(r'\D', '', xml_year)

                    collection = ""
                    type = ""
                    # if 'AVE' in record:
                    #     if 'm' in record['AVE']:
                    #         collection = record['AVE']['m']
                    #         type = 'temp'
                    url = ""
                    if collection != "" and xml_url != "":
                        url =  xml_url + "|" + collection
                    elif collection != "":
                        url = collection
                    elif xml_url != "":
                        url = xml_url

                    electronic_h_bool = False
                    if 'h' in record['245']:
                        if 'electronic resource' in record['245']['h'] or 'Electronic resource' in record['245']['h']:
                            electronic_h_bool = True
                            type = 'electronic_245_h'


                    if '655' in record and "Electronic books".lower() in record['655']['a'].lower():
                        type = 'electronic_655'
                    # print(record)
                    # print("Type in function: " + type)

                    records.append([xml_mms_id, xml_url, a, b, xml_title, xml_title_dash, xml_author_100s, xml_author_700s, xml_year, collection, url, type, record])

    # print(records)
    return(records)

def link_check(row, link, broken_links, link_success_counter, link_broken_counter, check_type):
    link = re.sub(r'\'', '\\"', link)
    link = re.sub(r'([\[\]\{\}])', r'\\\1', link)
    # print(link)
    #link = quote(link)
    #link_check_result = os.system('linkchecker -q "' + link + '"')
    curl_result = os.system('curl -g -s -o nul "' + link + '"')
    success = 'Fail'
    # print("\n\n\n\ncurl result: " + str(curl_result) + "\n\n\n\n")
    if check_type == 'individual':
        if curl_result != 0:
            # print("Link check failure")
            link_broken_counter += 1
            row = row.set_value('Link Status', 'Fail')
            broken_links = broken_links.append(row)
        else:
            # print("Link check success")
            row = row.set_value('Link Status', 'Success')
            link_success_counter += 1
            success = 'Success'

    elif check_type == 'compare':
        if curl_result != 0:
            # print("Link check failure")
            link_broken_counter += 1
            # row = row.set_value('Link Status', 'Fail')
            broken_links = broken_links.append(row)
        else:
            # print("Link check success")
            # row = row.set_value('Link Status', 'Success')
            link_success_counter += 1
            success = 'Success'
    #print("Curl result: " + str(curl_result))




    return([broken_links, row, link_success_counter, link_broken_counter, success])

    # broken_links = return_list_links[0]
    # electronic_list.iloc[f] = return_list_links[1]
    # link_success_counter = return_list_links[2]
    # link_broken_counter = return_list_links[3]



def ebook_match(record, m, t, a, ac, yr, xm, xt, xtd, xa100s, xa700s, xyr, xurl, el_list, ebook_match_on_list_counter, ebook_for_physical_counter, y, course_df, ebooks_to_add, ebooks_we_need, match_type, ebook_match_on_list, broken_links, link, link_success_counter, link_broken_counter, run_link_check):
    #ebook_match(bib_record, mms_id, title, author, author_contributor, year, xml_mms_id, xml_title, xml_title_dash, xml_author_100s, xml_author_700s, xml_year, xml_url, electronic_list, ebook_match_on_list_counter, ebook_for_physical_counter, i, electronic_list, ebooks_to_add, ebooks_we_need, "physical", ebook_match_on_list, broken_links, link, link_success_counter, link_broken_counter, run_link_check)
    master_match_type = ""
    loop_match_type = ""
    success = False
    # if run_link_check == "Yes":
    # print("\n\n\n\n" + str(course_df.iloc[y]) + "\n\n\n\n")

    # print("Source title in function:           " + t)
    # print("XML title in function:              " + xt)
    # print("Source author in function:          " + a)
    # print("Source contributor in function:     " + ac)
    # print("XML personal author in function:    " + xa100s)
    # print("XML corporate author in function:   " + xa700s)
    # print("Source year in function:            " + yr)
    # print("XML year in function:               " + xyr)
    if m == xm and match_type == "physical":
        if ((str(t) == str(xt) or t == xtd) and (xa100s == a or xa100s in a or a in xa100s or xa100s == ac or xa100s in ac or ac in xa100s or xa700s == a or xa700s in a or a in xa700s or ac == xa700s or xa700s in ac or ac in xa700s) and str(yr) == str(xyr)):
            #print("Full match\/Ebook match on record for " + t + " with " + xt)
            check_type = 'compare'
            return_list_links = link_check(course_df.iloc[y], xurl, broken_links, link_success_counter, link_broken_counter, check_type)

            broken_links = return_list_links[0]
            #course_df.iloc[y] = return_list_links[1]
            # print("\n\n\n\n" + str(course_df.iloc[y]['Link Status']) + "\n\n\n\n")
            # sys.exit()
            # print("\n\n\n\n new row in ebook match function with link checker: " + str(course_df.iloc[y]['Link Status']) + "\n\n\n\n")
            link_success_counter = return_list_links[2]
            link_broken_counter = return_list_links[3]
            success = return_list_links[4]
            loop_match_type = ave_loop(record)
            ebook_match_on_list_counter += 1
            #'Match on Course or Repo', 'Covid Collection or Permenant', 'Match on Year', 'Deleted Electronic Collection?'
            if loop_match_type == "non-Covid":
                course_df.iloc[y]['Covid Collection or Permenant'] = 'permanent'

            elif loop_match_type == 'Covid':
                #temporary_collections_portfolio_counter_on_course += 1
                course_df.iloc[y]['Covid Collection or Permenant'] = 'Covid'
            course_df.iloc[y]['Match on Course or Repo'] = 'course'
            course_df.iloc[y]['Match on Year'] = 'yes'
            course_df.iloc[y]['Deleted Electronic Record?'] = 'no'
            base_series = course_df.iloc[y]
            add_series = pd.Series({'Match MMS ID': xm, 'Match Title': xt, 'Match Author': xa100s + "; " + xa700s, 'Match Publication Year': xyr, 'Match URL or Collection': xurl, 'Match Link Status': success})
            series_to_add = base_series.append(add_series)
            ebook_match_on_list = ebook_match_on_list.append(course_df.iloc[y])
            success = True
        elif ((str(t) == str(xt) or t == xtd) and (xa100s == a or xa100s in a or a in xa100s or xa100s == ac or xa100s in ac or ac in xa100s or xa700s == a or xa700s in a or a in xa700s or ac == xa700s or xa700s in ac or ac in xa700s)):
            #print("Partial match\/Ebook match on record for " + t + " with " + xt)
            check_type = 'compare'
            return_list_links = link_check(course_df.iloc[y], xurl, broken_links, link_success_counter, link_broken_counter, check_type)

            broken_links = return_list_links[0]
            #course_df.iloc[y] = return_list_links[1]
            # print("\n\n\n\n" + str(course_df.iloc[y]['Link Status']) + "\n\n\n\n")
            # sys.exit()
            # print("\n\n\n\n new row in ebook match function with link checker: " + str(course_df.iloc[y]['Link Status']) + "\n\n\n\n")
            link_success_counter = return_list_links[2]
            link_broken_counter = return_list_links[3]
            success = return_list_links[4]
            loop_match_type = ave_loop(record)
            ebook_match_on_list_counter += 1
            #'Match on Course or Repo', 'Covid Collection or Permenant', 'Match on Year', 'Deleted Electronic Collection?'
            if loop_match_type == "non-Covid":
                course_df.iloc[y]['Covid Collection or Permenant'] = 'permanent'

            elif loop_match_type == 'Covid':
                #temporary_collections_portfolio_counter_on_course += 1
                course_df.iloc[y]['Covid Collection or Permenant'] = 'Covid'
            course_df.iloc[y]['Match on Course or Repo'] = 'course'
            course_df.iloc[y]['Match on Year'] = 'no'
            course_df.iloc[y]['Deleted Electronic Record?'] = 'no'
            base_series = course_df.iloc[y]
            add_series = pd.Series({'Match MMS ID': xm, 'Match Title': xt, 'Match Author': xa100s + "; " + xa700s, 'Match Publication Year': xyr, 'Match URL or Collection': xurl, 'Match Link Status': success})
            series_to_add = base_series.append(add_series)
            ebook_match_on_list = ebook_match_on_list.append(course_df.iloc[y])
            success = True
    elif m != xm:
        if (len(el_list[el_list['MMS Id'] == xm]) > 0) and match_type != "electronic":
            if ((str(t) == str(xt) or t == xtd) and (xa100s == a or xa100s in a or a in xa100s or xa100s == ac or xa100s in ac or ac in xa100s or xa700s == a or xa700s in a or a in xa700s or ac == xa700s or xa700s in ac or ac in xa700s) and str(yr) == str(xyr)):
                #print("Full match\/Ebook match on list for " + t + " with " + xt)
                check_type = 'compare'
                return_list_links = link_check(course_df.iloc[y], xurl, broken_links, link_success_counter, link_broken_counter, check_type)

                broken_links = return_list_links[0]
                #course_df.iloc[y] = return_list_links[1]
                # print("\n\n\n\n" + str(course_df.iloc[y]['Link Status']) + "\n\n\n\n")
                # sys.exit()
                # print("\n\n\n\n new row in ebook match function with link checker: " + str(course_df.iloc[y]['Link Status']) + "\n\n\n\n")
                link_success_counter = return_list_links[2]
                link_broken_counter = return_list_links[3]
                success = return_list_links[4]
                loop_match_type = ave_loop(record)
                ebook_match_on_list_counter += 1
                #'Match on Course or Repo', 'Covid Collection or Permenant', 'Match on Year', 'Deleted Electronic Collection?'
                if loop_match_type == "non-Covid":
                    course_df.iloc[y]['Covid Collection or Permenant'] = 'permanent'

                elif loop_match_type == 'Covid':
                    #temporary_collections_portfolio_counter_on_course += 1
                    course_df.iloc[y]['Covid Collection or Permenant'] = 'Covid'
                course_df.iloc[y]['Match on Course or Repo'] = 'course'
                course_df.iloc[y]['Match on Year'] = 'yes'
                course_df.iloc[y]['Deleted Electronic Record?'] = 'no'
                base_series = course_df.iloc[y]
                add_series = pd.Series({'Match MMS ID': xm, 'Match Title': xt, 'Match Author': xa100s + "; " + xa700s, 'Match Publication Year': xyr, 'Match URL or Collection': xurl, 'Match Link Status': success})
                series_to_add = base_series.append(add_series)
                ebook_match_on_list = ebook_match_on_list.append(course_df.iloc[y])
                success = True
            elif ((str(t) == str(xt) or t == xtd) and(xa100s == a or xa100s in a or a in xa100s or xa100s == ac or xa100s in ac or ac in xa100s or xa700s == a or xa700s in a or a in xa700s or ac == xa700s or xa700s in ac or ac in xa700s)):
                #print("Partial match\/Ebook match on list for " + t + " with " + xt)
                check_type = 'compare'
                return_list_links = link_check(course_df.iloc[y], xurl, broken_links, link_success_counter, link_broken_counter, check_type)

                broken_links = return_list_links[0]
                #course_df.iloc[y] = return_list_links[1]
                # print("\n\n\n\n" + str(course_df.iloc[y]['Link Status']) + "\n\n\n\n")
                # sys.exit()
                # print("\n\n\n\n new row in ebook match function with link checker: " + str(course_df.iloc[y]['Link Status']) + "\n\n\n\n")
                link_success_counter = return_list_links[2]
                link_broken_counter = return_list_links[3]
                success = return_list_links[4]
                loop_match_type = ave_loop(record)
                ebook_match_on_list_counter += 1
                #'Match on Course or Repo', 'Covid Collection or Permenant', 'Match on Year', 'Deleted Electronic Collection?'
                if loop_match_type == "non-Covid":
                    course_df.iloc[y]['Covid Collection or Permenant'] = 'permanent'

                elif loop_match_type == 'Covid':
                    #temporary_collections_portfolio_counter_on_course += 1
                    course_df.iloc[y]['Covid Collection or Permenant'] = 'Covid'
                course_df.iloc[y]['Match on Course or Repo'] = 'course'
                course_df.iloc[y]['Match on Year'] = 'no'
                course_df.iloc[y]['Deleted Electronic Record?'] = 'no'
                base_series = course_df.iloc[y]
                add_series = pd.Series({'Match MMS ID': xm, 'Match Title': xt, 'Match Author': xa100s + "; " + xa700s, 'Match Publication Year': xyr, 'Match URL or Collection': xurl, 'Match Link Status': success})
                series_to_add = base_series.append(add_series)
                ebook_match_on_list = ebook_match_on_list.append(course_df.iloc[y])
                success = True
        #elif (len(el_list[el_list['MMS Id'] == xm]) > 0) and match_type == "electronic" :
        else:

            if ((str(t) == str(xt) or t == xtd) and (xa100s == a or xa100s in a or a in xa100s or xa100s == ac or xa100s in ac or ac in xa100s or xa700s == a or xa700s in a or a in xa700s or ac == xa700s or xa700s in ac or ac in xa700s) and str(yr) == str(xyr)):
                #print("Full match\/Ebook match in Repo " + t + " with " + xt)
                loop_match_type = ave_loop(record)
                check_type = 'individual'
                return_list_links = link_check(course_df.iloc[y], xurl, broken_links, link_success_counter, link_broken_counter, type)

                broken_links = return_list_links[0]
                course_df.iloc[y] = return_list_links[1]
                # print("\n\n\n\n" + str(course_df.iloc[y]['Link Status']) + "\n\n\n\n")
                # sys.exit()
                # print("\n\n\n\n new row in ebook match function with link checker: " + str(course_df.iloc[y]['Link Status']) + "\n\n\n\n")
                link_success_counter = return_list_links[2]
                link_broken_counter = return_list_links[3]
                success = return_list_links[4]

                ebook_for_physical_counter += 1
                #'Match on Course or Repo', 'Covid Collection or Permenant', 'Match on Year', 'Deleted Electronic Collection?'
                if loop_match_type == "non-Covid":
                    course_df.iloc[y]['Covid Collection or Permenant'] = 'permanent'

                elif loop_match_type == 'Covid':
                    #temporary_collections_portfolio_counter_on_course += 1
                    course_df.iloc[y]['Covid Collection or Permenant'] = 'Covid'
                course_df.iloc[y]['Match on Course or Repo'] = 'repo'
                course_df.iloc[y]['Match on Year'] = 'yes'
                print(course_df.iloc[y])
                #course_df.iloc[y]['Deleted Electronic Record?'] = 'no'
                base_series = course_df.iloc[y]
                add_series = pd.Series({'Match MMS ID': xm, 'Match Title': xt, 'Match Author': xa100s + "; " + xa700s, 'Match Publication Year': xyr, 'Match URL or Collection': xurl, 'Match Link Status': success})
                series_to_add = base_series.append(add_series)

                ebooks_to_add = ebooks_to_add.append(series_to_add, ignore_index=True)

                success = True
            elif ((str(t) == str(xt) or t == xtd) and (xa100s == a or xa100s in a or a in xa100s or xa100s == ac or xa100s in ac or ac in xa100s or xa700s == a or xa700s in a or a in xa700s or ac == xa700s or xa700s in ac or ac in xa700s)):
                #print("Partial match\/Ebook match in Repo " + t + " with " + xt)
                loop_match_type = ave_loop(record)
                check_type = 'compare'
                return_list_links = link_check(course_df.iloc[y], xurl, broken_links, link_success_counter, link_broken_counter, check_type)

                broken_links = return_list_links[0]
                #course_df.iloc[y] = return_list_links[1]
                # print("\n\n\n\n" + str(course_df.iloc[y]['Link Status']) + "\n\n\n\n")
                # sys.exit()
                # print("\n\n\n\n new row in ebook match function with link checker: " + str(course_df.iloc[y]['Link Status']) + "\n\n\n\n")
                link_success_counter = return_list_links[2]
                link_broken_counter = return_list_links[3]
                success = return_list_links[4]

                ebook_for_physical_counter += 1
                #'Match on Course or Repo', 'Covid Collection or Permenant', 'Match on Year', 'Deleted Electronic Collection?'
                if loop_match_type == "non-Covid":
                    course_df.iloc[y]['Covid Collection or Permenant'] = 'permanent'

                elif loop_match_type == 'Covid':
                    #temporary_collections_portfolio_counter_on_course += 1
                    course_df.iloc[y]['Covid Collection or Permenant'] = 'Covid'
                course_df.iloc[y]['Match on Course or Repo'] = 'repo'
                course_df.iloc[y]['Match on Year'] = 'no'
                print(course_df.iloc[y])
                #course_df.iloc[y]['Deleted Electronic Record?'] = 'no'
                base_series = course_df.iloc[y]
                add_series = pd.Series({'Match MMS ID': xm, 'Match Title': xt, 'Match Author': xa100s + "; " + xa700s, 'Match Publication Year': xyr, 'Match URL or Collection': xurl, 'Match Link Status': success})
                series_to_add = base_series.append(add_series)
                # print(series_to_add)
                ebooks_to_add = ebooks_to_add.append(series_to_add, ignore_index=True)

                success = True
    return([success, [ebook_match_on_list_counter, ebook_for_physical_counter, link_success_counter, link_broken_counter], [ebooks_to_add, ebook_match_on_list, broken_links]])


oDir = "./Output"
if not os.path.isdir(oDir) or not os.path.exists(oDir):
    os.makedirs(oDir)

semester = input("Enter the course code prefix for the semester you'd like analyze: ")



chosen_proc_dept_integer = input("\n\nWhat library's reserve lists do you want to analyze.  (blank for all)?\n\n      1) Ginn\n      2) Hirsh\n      3) Music\n      4) SMFA\n      5) Tisch\n      6) Vet\n\nSelect an option: ")

chosen_proc_dept_integer = str(chosen_proc_dept_integer)
chosen_proc_dept = ""
if chosen_proc_dept_integer == "1":
    chosen_proc_dept = "Ginn Reserves"
elif chosen_proc_dept_integer == "2":
    chosen_proc_dept = "Hirsh Health Sciences Reserves"
elif chosen_proc_dept_integer == "3":
    chosen_proc_dept = "Music Reserves"
elif chosen_proc_dept_integer == "4":
    chosen_proc_dept = "SMFA Reserves"
elif chosen_proc_dept_integer == "5":
    chosen_proc_dept = "Tisch Reserves"
elif chosen_proc_dept_integer == "6":
    chosen_proc_dept = "Vet Reserves"
elif chosen_proc_dept_integer == "":
    chosen_proc_dept == ""
else:
    print("You made an invalid selection")
    exit(1)




reserves_filename = askopenfilename(title = "Select reserves filename")

start = datetime.datetime.now()

reserves_df = pd.read_excel(reserves_filename, dtype={'MMS Id': 'str', 'Publication Date': 'str', 'Title (Normalized)': 'str', 'Reading List Id': 'str', 'Citation Id': 'str'})

reserves_df_total = reserves_df.copy()


pd.set_option('display.max_columns', None)


reserves_df_total['Semester'] = reserves_df_total['Course Code'].apply(lambda x: re.sub(r'^(\d{4}).+', r'\1', x))
reserves_df_selected = reserves_df_total[reserves_df_total['Semester'] == semester]

reserves_df_selected = reserves_df_selected.fillna("")
for column in reserves_df_selected.columns:
    reserves_df_selected[column] = reserves_df_selected[column].apply(lambda x: str(x))
    # print(column)
    reserves_df_selected[column] = reserves_df_selected[column].apply(lambda x: re.sub(r'[\(\)]', '', x))

reserves_df_selected = reserves_df_selected.sort_values(by=['Processing Department', 'Course Code'])

column_list = reserves_df_selected.columns
proc_dept_df_list = []
proc_dept_list = []

if chosen_proc_dept == "Hirsh Health Sciences Reserves":
    hhsl_proc_dept_df = reserves_df_selected[reserves_df_selected['Processing Department'] == 'Hirsh Health Sciences Reserves']
    hhsl_proc_dept_df.append(reserves_df_selected[reserves_df_selected['Processing Department'] == 'Hirsh Reserves'])
    hhsl_proc_dept = "All Hirsh Reserves"
    proc_dept_df_list.append(hhsl_proc_dept_df)
    proc_dept_list.append(hhsl_proc_dept)

elif chosen_proc_dept != "":
    proc_dept_df_list.append(reserves_df_selected[reserves_df_selected['Processing Department'] == chosen_proc_dept])
    proc_dept_list.append(chosen_proc_dept)

else:
    for proc_dept, proc_dept_df in reserves_df_selected.groupby('Processing Department'):
        proc_dept_df_list.append(proc_dept_df)
        proc_dept_list.append(proc_dept)



d = 0


for dept in proc_dept_df_list:

    processing_dept = proc_dept_list[d]
    if processing_dept == 'Hirsh Reserves' or processing_dept == 'Hirsh Health Sciences Reserves':
        processing_dept = "All Hirsh Reserves"
    date = datetime.datetime.now().strftime("%m-%d-%Y")
    filename = oDir + '/' + processing_dept + ' - ' + date + '.xlsx'


    wb = openpyxl.Workbook()
    wb.save(filename)
    workbook = load_workbook(filename)
    writer = pd.ExcelWriter(filename, engine='openpyxl')
    writer.book = workbook



    course_df_list = []
    # reserves_df_selected = reserves_df_selected[reserves_df_selected['Processing Department'] == 'Hirsh Health Sciences Reserves']
    for course, course_df in proc_dept_df_list[d].groupby('Course Code'):
        # course_df['Link Status'] = 'Blank'
        #output_cols.extend(['Link Status', 'Match on Course or Repo', 'Covid Collection or Permenant', 'Match on Year', 'Deleted Electronic Record?', 'Match MMS ID', 'Match Title', 'Match Author', 'Match Publication Year', 'Match URL or Collection'])
        cols = course_df.columns.tolist()
        cols.insert(0, cols.pop(cols.index('Processing Department')))
        cols.insert(1, cols.pop(cols.index('Course Code')))
        cols.insert(2, cols.pop(cols.index('Course Name')))
        course_df = course_df.reindex(columns=[*course_df.columns.tolist(), 'Link Status', 'Match on Course or Repo', 'Covid Collection or Permenant', 'Match on Year', 'Deleted Electronic Record?'], fill_value="")


        course_df_list.append(course_df)




    ## testing



    x = 0
    col_list = reserves_df_selected.columns.tolist()
    # col_list_links = col_list.extend(['Link Status'])
    output_cols = reserves_df_selected.columns.tolist()

    # output_cols.extend(['Link Status'])


    output_cols.extend(['Link Status', 'Match on Course or Repo', 'Covid Collection or Permenant', 'Match on Year', 'Deleted Electronic Record?', 'Match MMS ID', 'Match Title', 'Match Author', 'Match Publication Year', 'Match URL or Collection'])
    counts_df = pd.DataFrame(columns=['Processing Department', 'Physical Books on Course', 'No Electronic Version for Physical Book', 'Electronic - In Collection - Add to Course'])
    ebooks_to_add = pd.DataFrame(columns=output_cols)
    print(ebooks_to_add)

    #ebooks_to_add_different_year = pd.DataFrame(columns=output_cols)

    ebooks_standalone_on_list = pd.DataFrame(columns=col_list)
    ebooks_we_need = pd.DataFrame(columns=col_list)
    ebook_match_on_list = pd.DataFrame(columns=output_cols)
    #ebook_match_on_list_without_year = pd.DataFrame(columns=output_cols)
    non_repository_citation_matches = pd.DataFrame(columns=col_list)
    non_repository_citation_no_match = pd.DataFrame(columns=col_list)
    broken_links = pd.DataFrame(columns=col_list)



    sru_url_prefix = "https://tufts.alma.exlibrisgroup.com/view/sru/01TUN_INST?version=1.2&operation=searchRetrieve&recordSchema=marcxml&query="


    #covid_e_books_df = pd.DataFrame(columns=output_cols)

    #covid_e_books_near_match_df = pd.DataFrame(columns=output_cols)
    while x < len(course_df_list):

        course_name = course_df_list[x]['Course Name']

        # if x >= 5:
        #     break
        y = 0
        #temporary_collections_portfolio_counter = 0
        #temporary_collections_portfolio_counter_on_course = 0
        #temporary_collections_counter_on_course_near_match = 0
        #temporary_collections_portfolio_counter_near_match = 0
        non_repository_citation_counter = 0
        non_repository_citation_counter_match = 0
        electronic_record_deleted_counter = 0
        ebook_counter = 0
        ebook_standalone_counter = 0
        ebook_for_physical_counter = 0
        no_match_electronic_counter = 0
        #different_year_ebook_for_physical_counter = 0
        #ebook_match_on_list_counter_without_year = 0
        link_success_counter = 0
        link_broken_counter = 0

        ebook_match_on_list_counter = 0
        #ebook_match_on_list_counter_without_year = 0
        #ebook_on_list_counter = 0
        #temporary_collection_ebook_on_list_counter = 0
        no_match_physical_counter = 0
        number_of_books_on_course = len(course_df_list[x])


        date = datetime.datetime.now().strftime("%m/%d/%Y")

        #counts_file = open(oDir + '/Counts - ' + str(course_name) + ' - ' + date + '.txt', 'w+')
        counts_per_course_list = []

        # print("Items on course: " + str(len(course_df_list[x])))
        course_ebook_count = 0
        course_code = course_df_list[x].iloc[0]['Course Code']
        course_name = course_df_list[x].iloc[0]['Course Name']

        print(course_code + "\n" + course_name + "\n")


        #physical_list = course_df_list[x][course_df_list[x]['Resource Type'] == 'Book - Physical']
        physical_list = course_df_list[x][course_df_list[x]['Resource Type'].str.match(r'(.*Physical.*)')==True]


        #electronic_list = course_df_list[x][course_df_list[x]['Resource Type'] == 'Book - Electronic']
        electronic_list = course_df_list[x][course_df_list[x]['Resource Type'].str.match(r'(.*Electronic.*)')==True]

        #match_output_cols.extend(['Link Status', 'Match on Course or Repo', 'Covid Collection or Permenant', 'Match on Year', 'Deleted Electronic Record?', 'Match MMS ID', 'Match Title', 'Match Author', 'Match Publication Year', 'Match URL or Collection'])

        electronic_list['Deleted Record?'] = ""



        non_repository_citation_list = course_df_list[x][course_df_list[x]['Is Repository Citation'] == 'No']

        number_of_physical_books_on_course = len(physical_list)



        # print(physical_list)
        if len(physical_list) > 0:


            while y < len(physical_list):
                print("Length of physical list: " + str(len(physical_list)))
                print("y: "                       + str(y))
                match = False
                course_df = physical_list.copy()

                course_df['Link Status'] = "Empty"



                title = course_df.iloc[y]['Title (Normalized)']
                title_for_query = '\"' + re.sub(r'\s', '%20', title) + '\"'
                author = ""
                author_contributor = ""
                # print(str(author) + "Type: " + str(type(author)))
                if course_df.iloc[y]['Author'] != "" and course_df.iloc[y]['Author'] is not np.nan:
                    #author = re.sub(r'([^,;]+,\s+[^,;],*\s*[^;,]*).+', r'\1', course_df.iloc[y]['Author'])
                    author = re.sub(r'(,\sauthor\.?|,\scontributor\.?|\scontributor\.?|\sauthor\.?)', '', course_df.iloc[y]['Author'])
                    author = author.lower()
                elif course_df.iloc[y]['Author (contributor)'] != "" and course_df.iloc[y]['Author (contributor)'] is not np.nan:
                    author_contributor = re.sub(r'(,\sauthor|,\scontributor|\scontributor|\sauthor)', '', course_df.iloc[y]['Author (contributor)'])
                    author_contributor = author_contributor.lower()
                #author_for_query = '\"' + re.sub(r'\s', '%20', author) + '\"'
                year = course_df.iloc[y]['Publication Date']
                year = re.sub(r'\D', '', year)

                mms_id = course_df.iloc[y]['MMS Id']


                # print("Length of Electonic List: " + str(len(electronic_list)) + "\n")

                print(course_df.iloc[y]['Publisher'])
                if 'ProQuest' in course_df.iloc[y]['Publisher'] and 'Theses' in course_df.iloc[y]['Publisher']:
                    print("\n\n\n\n\nProQuest Match\n\n\n\n\n")
                    ebook_match_on_list_counter += 1
                    course_df.iloc[y]['Match on Course or Repo'] = 'course'
                    course_df.iloc[y]['Covid Collection or Permanent'] = 'permanent'
                    course_df.iloc[y]['Match on Year'] = 'yes'

                    proquest_citation_id = course_df.iloc[y]['Citation Id']
                    proquest_reading_list_id = course_df.iloc[y]['Reading List Id']
                    proquest_course_id = course_df.iloc[y]['Course ID']
                    proquest_citation_url = "https://api-na.hosted.exlibrisgroup.com/almaws/v1/courses/" + str(proquest_course_id) + "/reading-lists/" + str(proquest_reading_list_id) + "/citations/" + str(proquest_citation_id) + "?apikey=l7xxde379ecb50e14de0959be6c41c1f6888&format=json"

                    proquest_citation_record = requests.get(proquest_citation_url).text

                    proquest_citation_record = json.loads(proquest_citation_record)
                    open_url_link = proquest_citation_record['open_url']
                    match = True
                    #link = course_df.iloc[y]['Citation Source']
                    check_type = 'individual'
                    return_list_links = link_check(course_df.iloc[y], open_url_link, broken_links, link_success_counter, link_broken_counter, check_type)


                    broken_links = return_list_links[0]
                    course_df.iloc[y] = return_list_links[1]
                    link_success_counter = return_list_links[2]
                    link_broken_counter = return_list_links[3]

                    ebook_match_on_list = ebook_match_on_list.append(course_df.iloc[y])
                    ebook_counter += 1
                    y += 1
                    continue

                if len(electronic_list) >= 1:


                    electronic_list['Author'] = electronic_list['Author'].apply(lambda x: re.sub(r'(,\sauthor|,\scontributor|\scontributor|\sauthor)', '', x).lower())
                    electronic_list['Author (contributor)'] = electronic_list['Author (contributor)'].apply(lambda x: re.sub(r'(,\sauthor|,\scontributor|\scontributor|\sauthor)', '', x).lower())
                    f = 0


                    for f in range(0, len(electronic_list)):
                        link = electronic_list.iloc[f]['Citation Source']
                        electronic_mms_id = electronic_list.iloc[f]['MMS Id']
                        electronic_record_result = requests.get(sru_url_prefix + 'alma.mms_id=' + electronic_mms_id)
                        deleted_record = ''
                        if "<errorsExist>true</errorsExist>" in str(electronic_record_result.content) or "<numberOfRecords>0</numberOfRecords>" in str(electronic_record_result.content):
                            deleted_record = 'yes'
                        else:
                            deleted_record = 'no'

                        if (electronic_list.iloc[f]['Title (Normalized)'] == title) and (electronic_list.iloc[f]['Author (contributor)'] in author or author in electronic_list.iloc[f]['Author (contributor)'] or electronic_list.iloc[f]['Author'] in author or author in electronic_list.iloc[f]['Author']) and (electronic_list.iloc[f]['Publication Date'] == year):
                            ebook_match_on_list_counter += 1
                            match = True
                            check_type = "individual"
                            return_list_links = link_check(electronic_list.iloc[f], link, broken_links, link_success_counter, link_broken_counter, check_type)

                            broken_links = return_list_links[0]
                            electronic_list.iloc[f] = return_list_links[1]
                            link_success_counter = return_list_links[2]
                            link_broken_counter = return_list_links[3]
    #output_cols.extend(['Match MMS ID', 'Match Title', 'Match Author', 'Match Publication Year', 'Match URL or Collection',
    #'Match on Course or Repo', 'Match in Repo?', 'Covid Collection?', 'Different Year?', 'Deleted Electronic Collection?'])
                            #'Match on Course or Repo', 'Covid Collection or Permenant', 'Match on Year', 'Deleted Electronic Record?'
                            electronic_list.iloc[f]['Match on Course or Repo'] = 'course'
                            electronic_list.iloc[f]['Covid Collection or Permanent'] = 'permanent'
                            electronic_list.iloc[f]['Match on Year'] = 'yes'
                            electronic_list.iloc[f]['Deleted Electronic Record?'] = deleted_record



                            ebook_match_on_list = ebook_match_on_list.append(electronic_list.iloc[f])


                            electronic_list = electronic_list.drop(electronic_list.index(f))
                            ebook_counter += 1


                            break

                        elif (electronic_list.iloc[f]['Title (Normalized)'] == title) and (electronic_list.iloc[f]['Author (contributor)'] in author or author in electronic_list.iloc[f]['Author (contributor)'] or electronic_list.iloc[f]['Author'] in author or author in electronic_list.iloc[f]['Author']):
                            ebook_match_on_list_counter += 1
                            check_type = "individual"
                            return_list_links = link_check(electronic_list.iloc[f], link, broken_links, link_success_counter, link_broken_counter, check_type)

                            broken_links = return_list_links[0]
                            electronic_list.iloc[f] = return_list_links[1]
                            link_success_counter = return_list_links[2]
                            link_broken_counter = return_list_links[3]
                            electronic_list.iloc[f]['Match on Course or Repo'] = 'course'
                            electronic_list.iloc[f]['Covid Collection or Permanent'] = 'permanent'
                            electronic_list.iloc[f]['Match on Year'] = 'no'
                            electronic_list.iloc[f]['Deleted Electronic Record?'] = deleted_record
                            ebook_match_on_list = ebook_match_on_list.append(electronic_list.iloc[f])
                            quasi_match_bool = True
                            match = True

                            electronic_list = electronic_list.drop(electronic_list.index(f))
                            ebook_counter += 1
                            break
                        f += 1



                if 'http' in course_df.iloc[y]['Citation Source']:
                    non_repository_citation_counter_match += 1
                    non_repository_citation_counter += 1
                    match = True
                    link = course_df.iloc[y]['Citation Source']
                    check_type = 'individual'
                    return_list_links = link_check(course_df.iloc[y], link, broken_links, link_success_counter, link_broken_counter, check_type)


                    broken_links = return_list_links[0]
                    course_df.iloc[y] = return_list_links[1]
                    link_success_counter = return_list_links[2]
                    link_broken_counter = return_list_links[3]


                    non_repository_citation_matches = non_repository_citation_matches.append(course_df.iloc[y], ignore_index=True)
                    # g += 1
                    y += 1
                    continue

                elif len(non_repository_citation_list) > 0:
                    g = 0
                    while g < len(non_repository_citation_list):
                        citation_id = non_repository_citation_list.iloc[g]['Citation Id']
                        reading_list_id = non_repository_citation_list.iloc[g]['Reading List Id']
                        course_record_url = "https://api-na.hosted.exlibrisgroup.com/almaws/v1/courses?apikey=l7xxde379ecb50e14de0959be6c41c1f6888&q=code~{" + course_code + "}&format=json"

                        # course_record = requests.get(course_record_url).json()
                        course_record = requests.get(course_record_url).text
                        # print(course_record)
                        course_record = json.loads(course_record)

                        course_id = course_record['course'][0]['id']


                        citation_url = "https://api-na.hosted.exlibrisgroup.com/almaws/v1/courses/" + str(course_id) + "/reading-lists/" + str(reading_list_id) + "/citations/" + str(citation_id) + "?apikey=l7xxde379ecb50e14de0959be6c41c1f6888&format=json"

                        citation_record = requests.get(citation_url).text




                        citation_record = json.loads(citation_record)
                        # print("\n\n\nCitation record: " + str(citation_record) + "\n\n\n\n")
                        # print("Citation Record: \n" + json.dumps(citation_record))
                        citation_title = citation_record['metadata']['title']
                        citation_author = citation_record['metadata']['author']
                        citation_source = citation_record['metadata']['source']

                        if citation_title is None or citation_author is None:
                            g += 1
                            continue
                        citation_title = citation_title.lower()
                        citation_title = re.sub(r'\s\/$', '', citation_title)
                        citation_title = re.sub(r'\'', ' ', citation_title)

                        citation_title = re.sub(r'\.$', '', citation_title)
                        citation_title = re.sub(r'^(the\s|a\s)', '', citation_title)
                        citation_title = re.sub(r'\s{2,}', ' ', citation_title)


                        citation_author = re.sub(r'(,\sauthor\.?|,\scontributor\.?|\scontributor\.?|\sauthor\.?)', '', citation_author)
                        citation_author = citation_author.lower()

                        citation_author = re.sub(r'(,\sauthor\.?|,\scontributor\.?|\scontributor\.?|\sauthor\.?)', '', citation_author)
                        citation_author = citation_author.lower()

                        # print('Physical source title:             ' + str(title))
                        # print('Physical course author:            ' + str(author))
                        # print('Physical source author\/cont:      ' + str(author_contributor))
                        # print('Citation title:                    ' + str(citation_title))
                        # print('Citation author:                   ' + str(citation_author))
                        # print('\n\n\n')
                        if (citation_title == title) and (citation_author in author or author in citation_author or citation_author in author_contributor or author_contributor in citation_author):
                            link = non_repository_citation_list.iloc[g]['Citation Source']
                            check_type = 'individual'
                            return_list_links = link_check(non_repository_citation_list.iloc[g], link, broken_links, link_success_counter, link_broken_counter, check_type)

                            broken_links = return_list_links[0]
                            non_repository_citation_list.iloc[g] = return_list_links[1]
                            link_success_counter = return_list_links[2]
                            link_broken_counter = return_list_links[3]

                            non_repository_citation_counter_match += 1
                            non_repository_citation_counter += 1
                            match = True

                            non_repository_citation_list.iloc[f]['Match on Course or Repo'] = 'course'

                            non_repository_citation_list.iloc[f]['Match on Year'] = 'no'

                            series = non_repository_citation_list[g]
                            add_series = pd.Series({'Citation Title': citation_title, 'Citation Author': citation_author, 'Citation Source': citation_source})
                            series_to_add = base_series.append(add_series)
                            non_repository_citation_matches = non_repository_citation_matches.append(series_to_add, ignore_index=True)

                            non_repository_citation_list = non_repository_citation_list.drop(non_repository_citation_list.index(g))
                            g += 1
                            break


                        g += 1



                    # else:
                #
                if match == True:
                    y += 1
                    continue



                return_list_1 = get_xml(title_for_query, sru_url_prefix)

                #return([xml_mms_id, xml_url, a, b, xml_title, xml_title_dash, xml_author, xml_year, collection, url, type])

                for result in return_list_1:

                    xml_mms_id = result[0]
                    xml_url = result[1]
                    a = result[2]
                    b = result[3]
                    xml_title = result[4]
                    xml_title_dash = result[5]
                    xml_author_100s = result[6]
                    xml_author_700s = result[7]
                    xml_year = result[8]
                    collection = result[9]
                    url = result[10]
                    type = result[11]
                    bib_record = result[12]



                    z = 0
                    # if bib_record is None:
                    #     continue
                    if 'AVE' in bib_record or 'electronic_655' in type:
                        # link = bib_record[0]['856']['a']
                        # return_list_links = link_check(course_df.iloc[y], link, broken_links, link_success_counter, link_broken_counter)
                        #
                        # broken_links = return_list_links[0]
                        # non_repository_citation_list.iloc[g] = return_list_links[1]
                        # link_success_counter = return_list_links[2]
                        # link_broken_counter = return_list_links[3]

                        run_link_check = "Yes"
                        return_list = ebook_match(bib_record, mms_id, title, author, author_contributor, year, xml_mms_id, xml_title, xml_title_dash, xml_author_100s, xml_author_700s, xml_year, xml_url, electronic_list, ebook_match_on_list_counter, ebook_for_physical_counter, y, course_df, ebooks_to_add, ebooks_we_need, "physical", ebook_match_on_list, broken_links, link, link_success_counter, link_broken_counter, run_link_check)
                        match = return_list[0]
                        counts_return = return_list[1]
                        dataframe_return = return_list[2]

            #return([success, [ebook_match_on_list_counter, ebook_for_physical_counter, link_success_counter, link_broken_counter], [ebooks_to_add, ebook_match_on_list, broken_links]])
                        #counts
                        ebook_match_on_list_counter = counts_return[0]
                        ebook_for_physical_counter = counts_return[1]
                        link_success_counter = counts_return[2]
                        link_broken_counter = counts_return[3]

                        #dataframes
                        #ebooks_to_add, ebook_match_on_list, broken_links
                        #course_df
                        ebooks_to_add = dataframe_return[0]
                        ebook_match_on_list = dataframe_return[1]
                        broken_links = dataframe_return[2]

                        if match == True:
                        	z += 1
                        	break



                        z += 1


                if not match:
                    # print("Defaulted into not found")
                    # base_series = course_df.iloc[y]
                        #add_series = pd.Series({'Course Code': course_code, 'Course Name': course_name, 'MMS ID': mms_id, 'Title': title, 'Author': author, 'Publication Year': year, 'Match MMS ID': xml_mms_id, 'Match Title': xml_title, 'Match Author': xml_author, 'Match Publication Year': xml_year, 'Match URL or Collection': url})
                        # series_to_add = base_series.append(add_series)
                        # series_to_add = base_series.append({'Match MMS ID': xml_mms_id, 'Match Title': xml_title, 'Match Author': xml_author, 'Match Publication Year': xml_year, 'Match URL or Collection': url})
                    ebooks_we_need = ebooks_we_need.append(course_df.iloc[y], ignore_index=True)
                    #ebooks_we_need = ebooks_we_need.append(course_df.iloc[y], ignore_index=True)
                    no_match_physical_counter += 1
                    y += 1
                    break

                        #print(ebooks_on_this_course)
                y += 1
        i = 0
        #
        # print("\n\n\n\n\n\n" + str(electronic_list) + "\n\n\n\n\n\n")
        while i < len(electronic_list):
            ebook_counter += 1
            # print("Length of Electonic List (lower): " + str(len(electronic_list)) + "\n")
            # print("Ebook counter:                    " + str(ebook_counter))
            # print("i:                                " + str(i))
            # print(str(electronic_list.iloc) + "\n")
            ebook_match_non_covid = False
            ebook_match_covid = False
            electronic_mms_id = electronic_list.iloc[i]['MMS Id']
            # print(electronic_mms_id)
            electronic_record_result = requests.get(sru_url_prefix + 'alma.mms_id=' + electronic_mms_id)
            # print(electronic_record_result.content)
            link = electronic_list.iloc[i]['Citation Source']
            check_type = 'individual'
            return_list_links = link_check(electronic_list.iloc[i], link, broken_links, link_success_counter, link_broken_counter, check_type)

            broken_links = return_list_links[0]
            electronic_list.iloc[i] = return_list_links[1]
            link_success_counter = return_list_links[2]
            link_broken_counter = return_list_links[3]
            if "<errorsExist>true</errorsExist>" in str(electronic_record_result.content) or "<numberOfRecords>0</numberOfRecords>" in str(electronic_record_result.content):
                electronic_record_deleted_counter += 1
                electronic_list.iloc[i]['Deleted Record?'] = 'Yes'
                # print(str(electronic_list) + "\n")
                # print("Errors in:" + str(electronic_list.iloc[i]))
                title_2 = electronic_list.iloc[i]['Title (Normalized)']
                title_for_query_2 = '\"' + re.sub(r'\s', '%20', title_2) + '\"'
                author_2 = electronic_list.iloc[i]['Author'].lower()
                author_contributor_2 = electronic_list.iloc[i]['Author (contributor)'].lower()
                # print(str(author) + "Type: " + str(type(author)))
                # if electronic_list.iloc[i]['Author'] != "" and electronic_list.iloc[i]['Author'] is not np.nan:
                #     author_2 = electronic_list.iloc[i]['Author']
                #     #author = re.sub(r'([^,;]+,\s+[^,;],*\s*[^;,]*).+', r'\1', course_df.iloc[y]['Author'])
                #     #author_2 = re.sub(r'(,\sauthor\.?|,\scontributor\.?|\scontributor\.?|\sauthor\.?)', '', author_2)
                #     author_2 = author.lower()
                #     # author_2 = re.sub(r'\.$', '' , author_2)
                #
                # else:
                #     author_2 = ""
                found_match = False
                #if electronic_list.iloc[i]['Author (contributor)'] != "" and electronic_list.iloc[i]['Author (contributor)'] is not np.nan:
                #     author_contributor_2 = electronic_list.iloc[i]['Author (contributor)']
                #     #author_contributor_2 = re.sub(r'(,\sauthor|,\scontributor|\scontributor|\sauthor)', '', author_contributor_2)
                #     author_contributor_2 = author_contributor_2.lower()
                #     # author_contributor_2 = re.sub(r'\.$', '' , author_contributor_2)
                #author_for_query = '\"' + re.sub(r'\s', '%20', author) + '\"'
                # print("Author 2: " + author_2)
                # print("Author Contributor: " + author_contributor_2)
                year_2 = electronic_list.iloc[i]['Publication Date']
                year_2 = re.sub(r'\D', '', year)

                mms_id_2 = electronic_list.iloc[i]['MMS Id']
                return_list_2 = get_xml(title_for_query_2, sru_url_prefix)

                for  item in return_list_2:
                    # print("Return list ...: " + str(item))
                    xml_mms_id = item[0]
                    xml_url = item[1]
                    a = item[2]
                    b = item[3]
                    xml_title = item[4]
                    xml_title_dash = item[5]
                    xml_author_100s = item[6]
                    xml_author_700s = item[7]
                    xml_year = item[8]
                    collection = item[9]
                    url = item[10]
                    type = item[11]
                    bib_record = item[12]
                    #
                    # print("Source electronic title:              " + title_2)
                    # print("XML electronic title:                 " + xml_title)
                    # print("Source electronic author:             " + author_2)
                    # print("Source electronic author/contributor: " + author_contributor_2)
                    # print("XML electronic personal author:       " + xml_author_100s)
                    # print("XML electronic corporate author:      " + xml_author_700s)
                    if 'AVE' in bib_record or 'electronic_655' in type:

                        #
                        # print("Source title:  " + title_2)
                        # print("Source author: " + author_2)
                        # print("Source author contributor: " + author_contributor_2)
                        # print("XML Title:     " + xml_title)
                        # # print("Source author: " + author)
                        # print("XML Author 100s:    " + xml_author_100s)
                        # print("XML Author 700s:    " + xml_author_700s   )
                        # print("MMS Id XML:    " + xml_mms_id)
                        match_type = "electronic"
                        run_link_check = "Yes"
                        print(ebooks_to_add)
                        #return_list_2 = ebook_match(bib_record, mms_id_2, title_2, author_2, author_contributor_2, year_2, xml_mms_id, xml_title, xml_title_dash, xml_author_100s, xml_author_700s, xml_year, xml_url, electronic_list, ebook_match_on_list_counter, ebook_match_on_list_counter_without_year, temporary_collections_portfolio_counter_on_course, temporary_collections_counter_on_course_near_match, ebook_for_physical_counter, different_year_ebook_for_physical_counter, temporary_collections_portfolio_counter, temporary_collections_portfolio_counter_near_match, i, electronic_list, ebooks_to_add, ebooks_to_add_different_year, ebooks_we_need, covid_e_books_df, covid_e_books_near_match_df, match_type, ebook_match_on_list, ebook_match_on_list_without_year, broken_links, link, link_success_counter, link_broken_counter, run_link_check)
                        return_list_2 = return_list = ebook_match(bib_record, mms_id_2, title_2, author_2, author_contributor_2, year_2, xml_mms_id, xml_title, xml_title_dash, xml_author_100s, xml_author_700s, xml_year, xml_url, electronic_list, ebook_match_on_list_counter, ebook_for_physical_counter, i, electronic_list, ebooks_to_add, ebooks_we_need, "physical", ebook_match_on_list, broken_links, link, link_success_counter, link_broken_counter, run_link_check)

                        match_2 = return_list_2[0]
                        counts_return_2 = return_list_2[1]
                        dataframe_return_2 = return_list_2[2]

                        #counts
                        ebook_match_on_list_counter = counts_return_2[0]
                        ebook_for_physical_counter = counts_return_2[1]
                        link_success_counter = counts_return_2[2]
                        link_broken_counter = counts_return_2[3]

                        #dataframes
                        #ebooks_to_add, ebook_match_on_list, broken_links
                        ebooks_to_add = dataframe_return_2[0]
                        ebook_match_on_list = dataframe_return_2[1]
                        broken_links = dataframe_return_2[2]

                        if match_2 == True:
                            found_match = True

                            break

                if found_match == False:
                    j = 0

                    if len(non_repository_citation_list) > 0:
                        while j < len(non_repository_citation_list):
                            citation_id = non_repository_citation_list.iloc[j]['Citation Id']
                            reading_list_id = non_repository_citation_list.iloc[j]['Reading List Id']
                            course_record_url = "https://api-na.hosted.exlibrisgroup.com/almaws/v1/courses?apikey=l7xxde379ecb50e14de0959be6c41c1f6888&q=code~{" + course_code + "}&format=json"

                            # course_record = requests.get(course_record_url).json()
                            course_record = requests.get(course_record_url).text
                            # print(course_record)
                            course_record = json.loads(course_record)

                            course_id = course_record['course'][0]['id']

                            citation_url = "https://api-na.hosted.exlibrisgroup.com/almaws/v1/courses/" + str(course_id) + "/reading-lists/" + str(reading_list_id) + "/citations/" + str(citation_id) + "?apikey=l7xxde379ecb50e14de0959be6c41c1f6888&format=json"

                            citation_record = requests.get(citation_url).text




                            citation_record = json.loads(citation_record)
                            # print("Citation Record: \n" + json.dumps(citation_record))
                            citation_title = citation_record['metadata']['title']
                            citation_author = citation_record['metadata']['author']
                            if citation_title is None:
                                citation_title = ""

                            if citation_author is None:
                                citation_author = ""
                            # print(json.dumps(citation_record))
                            citation_title = citation_title.lower()
                            citation_title = re.sub(r'\s\/$', '', citation_title)
                            citation_title = re.sub(r'\'', ' ', citation_title)

                            citation_title = re.sub(r'\.$', '', citation_title)
                            citation_title = re.sub(r'^(the\s|a\s)', '', citation_title)
                            citation_title = re.sub(r'\s{2,}', ' ', citation_title)


                            citation_author = re.sub(r'(,\sauthor\.?|,\scontributor\.?|\scontributor\.?|\sauthor\.?)', '', citation_author)
                            citation_author = citation_author.lower()

                            citation_author = re.sub(r'(,\sauthor\.?|,\scontributor\.?|\scontributor\.?|\sauthor\.?)', '', citation_author)
                            citation_author = citation_author.lower()
                            # print('Electronic source title:             ' + str(title_2))
                            # print('Electroniccourse author:            ' + str(author_2))
                            # print('Electronic source author\/cont:      ' + str(author_contributor_2))
                            # print('Citation title:                    ' + str(citation_title))
                            # print('Citation author:                   ' + str(citation_author))
                            # print('\n\n\n')
                            if (citation_title == title_2) and (citation_author in author_2 or author_2 in citation_author or citation_author in author_contributor_2 or author_contributor_2 in citation_author):
                                non_repository_citation_counter_match += 1
                                non_repository_citation_counter += 1
                                match = True
                                found_match = match
                                match_type = 'individual'
                                return_list_links = link_check(non_repository_citation_list.iloc[i], link, broken_links, link_success_counter, link_broken_counter, check_type)

                                broken_links = return_list_links[0]
                                non_repository_citation_list.iloc[i] = return_list_links[1]
                                link_success_counter = return_list_links[2]
                                link_broken_counter = return_list_links[3]

                                non_repository_citation_list.iloc[j]['Match on Course or Repo'] = 'repo'

                                non_repository_citation_list.iloc[j]['Match on Year'] = 'no'
                                series = non_repository_citation_list[j]



                                add_series = pd.Series({'Citation Title': citation_title, 'Citation Author': citation_author})
                                series_to_add = base_series.append(add_series)
                                non_repository_citation_matches = non_repository_citation_matches.append(series_to_add, ignore_index=True)

                                non_repository_citation_list = non_repository_citation_list.drop(non_repository_citation_list.index(j))
                                j += 1
                                break


                            j += 1


                        if found_match == False:


                            #
                            #
                            # print("Defaulted into not found")
                            # base_series = course_df.iloc[y]
                                #add_series = pd.Series({'Course Code': course_code, 'Course Name': course_name, 'MMS ID': mms_id, 'Title': title, 'Author': author, 'Publication Year': year, 'Match MMS ID': xml_mms_id, 'Match Title': xml_title, 'Match Author': xml_author, 'Match Publication Year': xml_year, 'Match URL or Collection': url})
                                # series_to_add = base_series.append(add_series)
                                # series_to_add = base_series.append({'Match MMS ID': xml_mms_id, 'Match Title': xml_title, 'Match Author': xml_author, 'Match Publication Year': xml_year, 'Match URL or Collection': url})
                            ebooks_we_need = ebooks_we_need.append(electronic_list.iloc[i], ignore_index=True)

                            no_match_electronic_counter += 1
                            # break

            else:
                #ebook_match_on_list_counter += 1
                ebooks_standalone_on_list = ebooks_standalone_on_list.append(electronic_list.iloc[i], ignore_index=True)
                ebook_standalone_counter += 1




            i += 1

            # else:

                # tree_e = et.ElementTree(et.fromstring(electronic_record_result.content))
                # root_e = tree_e.getroot()
                #
                #
                # # result_df = pd.DataFrame(columns=['MMS Id', 'Title', 'Author', 'Year'])
                #
                # for elem_e in root_e.iter():
                #
                #
                #     if re.match(r'.*record$', elem_e.tag):
                #
                #         bib_record_e = pym.parse_xml_to_array(io.StringIO(et.tostring(elem_e).decode('utf-8')))
                #         if 'AVE' in bib_record_e[0]:
                #
                #             match_type_e = ave_loop(bib_record_e[0])
                #
                #             if match_type_e == "non-Covid":
                #                 ebook_match_non_covid = True
                #
                #             elif match_type_e == 'Covid':
                #                 ebook_match_covid = True
                #
                # if ebook_match_non_covid == True:
                #     ebook_on_list_counter += 1
                # elif ebook_match_covid == True:
                #     temporary_collection_ebook_on_list_counter += 1

        h = 0
        if len(non_repository_citation_list) > 0:
            while h < len(non_repository_citation_list):
                non_repository_citation_counter += 1
                link = non_repository_citation_list.iloc[h]['Citation Source']
                check_type = 'individual'
                return_list_links = link_check(non_repository_citation_list.iloc[h], link, broken_links, link_success_counter, link_broken_counter, check_type)

                broken_links = return_list_links[0]
                non_repository_citation_list.iloc[h] = return_list_links[1]
                link_success_counter = return_list_links[2]
                link_broken_counter = return_list_links[3]
                #
                # match = True
                non_repository_citation_list.iloc[h]['Match on Course or Repo'] = 'repo'

                non_repository_citation_list.iloc[h]['Match on Year'] = 'no'
                series = non_repository_citation_list.iloc[h]
                #add_series = pd.Series({'Citation Title': citation_title, 'Citation Author': citation_author})
                #series_to_add = base_series.append(add_series)
                non_repository_citation_no_match = non_repository_citation_no_match.append(series, ignore_index=True)

                # non_repository_citation_list = non_repository_citation_list.drop(non_repository_citation_list.index(g))

                h += 1



        course_name = course_df_list[x].iloc[0]['Course Name']
        course_code = course_df_list[x].iloc[0]['Course Code']
        processing_dept = course_df_list[x].iloc[0]['Processing Department']


        #counts_file = open(oDir + '/Counts - ' + str(course_name) + ' - ' + date + '.txt', 'w+')

        #non_repository_citation_counter = 0
        #non_repository_citation_counter_match = 0
        #electronic_record_deleted_counter = 0
        #ebook_counter = 0
        #ebook_standalone_counter = 0
        #ebook_for_physical_counter = 0
        #no_match_electronic_counter = 0link_success_counter = 0
        #link_broken_counter = 0ebook_match_on_list_counter = 0no_match_physical_counter = 0
        counts_df = counts_df.append({'Processing Department': processing_dept, 'Course Name': course_name, 'Course Code': course_code, 'Citations on Course': number_of_books_on_course,  'Physical Books on Course' :number_of_physical_books_on_course, 'Electronic Books on Course': ebook_counter, 'Standalone Electronic on Course': ebook_standalone_counter, 'No Electronic Version for Physical Book': no_match_physical_counter, 'No Match Inactive Electronic with Citation': no_match_electronic_counter, 'Non-Repository - on Course': non_repository_citation_counter, 'Electronic - Match on Course': ebook_match_on_list_counter, 'Non-Repository Citation Matches': non_repository_citation_counter_match, 'Electronic - In Collection - Add to Course': ebook_for_physical_counter, 'Link Sucess': link_success_counter, 'Link Fail': link_broken_counter}, ignore_index=True)


        x += 1
    # This section moves the last two columns, course data, to the first two positions, and puts proc dept in beginning
    cols = counts_df.columns.tolist()
    cols = cols[-1:] + cols[:-1]
    counts_df = counts_df[cols]
    cols = cols[-1:] + cols[:-1]
    counts_df = counts_df[cols]
    cols.remove('Processing Department')
    cols.remove('Course Code')
    cols.remove('Course Name')
    cols.remove('Electronic Books on Course')
    cols.remove('Electronic - Match on Course')
    cols.remove('Non-Repository - on Course')
    cols.remove('Physical Books on Course')
    cols.remove('Citations on Course')
    cols.insert(0, 'Processing Department')
    cols.insert(1, 'Course Code')
    cols.insert(2, 'Course Name')
    cols.insert(3, 'Citations on Course')
    cols.insert(4, 'Physical Books on Course')
    cols.insert(5, 'Electronic Books on Course')
    cols.insert(6, 'Non-Repository - on Course')
    cols.insert(7, 'Electronic - Match on Course')


    counts_df = counts_df[cols]




    counts_df.to_excel(writer, sheet_name='Counts', index=False, engine='openpyxl')
    ebooks_to_add.to_excel(writer, sheet_name='InRepo', index=False, engine='openpyxl')
    ebooks_we_need.to_excel(writer, sheet_name='NeededNotInTufts', index=False, engine='openpyxl')
    non_repository_citation_no_match.to_excel(writer, sheet_name='Non-RepoStandAlone', index=False, engine='openpyxl')
    ebook_match_on_list.to_excel(writer, sheet_name='EbookMatch On List - Check URL', index=False, engine='openpyxl')
    ebooks_standalone_on_list.to_excel(writer, sheet_name='EbooksStandAlone-CheckURL', index=False, engine='openpyxl')
    broken_links.to_excel(writer, sheet_name='BrokenLinks', index=False, engine='openpyxl')
    ebooks_we_need_unique = ebooks_we_need.drop_duplicates(subset=['MMS Id'])



    ebooks_we_need_unique.to_excel(writer, sheet_name='UniquePrchsList', index=False, engine='openpyxl')




    for sheet in writer.sheets:
        e = 0
        for column in writer.sheets[sheet].iter_cols():
            writer.sheets[sheet].column_dimensions[get_column_letter(e + 1)].width = "20"
            e += 1
    f = 0

    # for sheet in writer.sheets:
    #     writer.sheets[sheet].freeze_panes = 'A2'
    #     for row in writer.sheets[sheet]:
    #         # print(row)
    #         if f == 0:
    #
    #             for cell in row:
    #                 #cell.style.alignment.wrap_text=True
    #                 cell.alignment = Alignment(wrap_text=True)
    #                 cell.font = Font(bold=True)
    #
    #         for cell in row:
    #             if cell.value == "Fail":
    #                 cell.font(color='red')
    #
    #         f += 1

    # sums
    counts_max_row = writer.book['Counts'].max_row

    for x in range(1, 21):
        writer.sheets['Counts'].cell(row = counts_max_row + 2, column = x).font = Font(bold=True)

    writer.sheets['Counts'].cell(row = counts_max_row + 2, column = 1).value = "Totals"
    writer.sheets['Counts'].cell(row = counts_max_row + 2, column = 4).value = '= SUM(D1:D' + str(counts_max_row) + ')'
    writer.sheets['Counts'].cell(row = counts_max_row + 2, column = 5).value = '= SUM(E1:E' + str(counts_max_row) + ')'
    writer.sheets['Counts'].cell(row = counts_max_row + 2, column = 6).value = '= SUM(F1:F' + str(counts_max_row) + ')'
    writer.sheets['Counts'].cell(row = counts_max_row + 2, column = 7).value = '= SUM(G1:G' + str(counts_max_row) + ')'
    writer.sheets['Counts'].cell(row = counts_max_row + 2, column = 8).value = '= SUM(H1:H' + str(counts_max_row) + ')'
    writer.sheets['Counts'].cell(row = counts_max_row + 2, column = 9).value = '= SUM(I1:I' + str(counts_max_row) + ')'
    writer.sheets['Counts'].cell(row = counts_max_row + 2, column = 10).value = '= SUM(J1:J' + str(counts_max_row) + ')'
    writer.sheets['Counts'].cell(row = counts_max_row + 2, column = 11).value = '= SUM(K1:K' + str(counts_max_row) + ')'
    writer.sheets['Counts'].cell(row = counts_max_row + 2, column = 12).value = '= SUM(L1:L' + str(counts_max_row) + ')'
    writer.sheets['Counts'].cell(row = counts_max_row + 2, column = 13).value = '= SUM(M1:M' + str(counts_max_row) + ')'
    writer.sheets['Counts'].cell(row = counts_max_row + 2, column = 14).value = '= SUM(N1:N' + str(counts_max_row) + ')'
    writer.sheets['Counts'].cell(row = counts_max_row + 2, column = 15).value = '= SUM(O1:O' + str(counts_max_row) + ')'


    #
    #
    first_sheet = workbook.get_sheet_by_name('Sheet')



    workbook.remove_sheet(first_sheet)
    writer.save()
    workbook.save(filename)
# link_check_result = os.system('linkchecker -q https://tufts.box.com/s/o574wls3hnrfirylv7hcejbh7it9awhq')
#
# if link_check_result == 0:
#     print("Link check success")
# else:
#     print("Link check failure")


    d += 1
print('\n\n\n')

end = datetime.datetime.now() - start
# test_file.close()

print("Execution time: " + str(end) + "\n")
