#!/usr/bin/env python3
import sys

import requests
import json
import os
import sys
import time
import csv
import re
import datetime

from tkinter.filedialog import askopenfilename

import pandas as pd
import numpy as np

import xml.etree.cElementTree as et
import pymarc as pym
import io
# from django.utils.encoding import smart_str

import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font

def ave_loop(record):
    perm_bool = False
    match_type = ""
    for ave in record.get_fields('AVE'):
        if 'm' in ave:
            if ('Covid-19'.lower() not in ave['m'].lower()):
                perm_bool = True
                match_type = "non-Covid"
                break
    if perm_bool == False:

        match_type = "Covid"

    return(match_type)

def ebook_match(record, m, t, a, yr, xm, xt, xtd, xa, xyr, el_list, ebook_match_on_list_counter, ebook_match_on_list_counter_without_year, temporary_collections_portfolio_counter_on_course, temporary_collections_counter_on_course_near_match, ebook_for_physical_counter, different_year_ebook_for_physical_counter, temporary_collections_portfolio_counter, temporary_collections_portfolio_counter_near_match, y, course_df, ebooks_to_add, ebooks_to_add_different_year, ebooks_we_need, covid_e_books_df, covid_e_books_near_match_df):

    master_match_type = ""
    loop_match_type = ""
    success = False

    if m == xm:
        if ((str(t) == str(xt) or t == xtd) and (xa == a or xa in a or a in xa) and str(yr) == str(xyr)):
            print(ebook_match_on_list_counter)
            loop_match_type = ave_loop(record)
            if loop_match_type == "non-Covid":
                ebook_match_on_list_counter += 1
            elif loop_match_type == 'Covid':
                temporary_collections_portfolio_counter_on_course += 1
            success = True
        elif ((str(t) == str(xt) or t == xtd) and (xa == a or xa in a or a in xa)):
            loop_match_type = ave_loop(record)
            if loop_match_type == "non-Covid":
                ebook_match_on_list_counter_without_year += 1
            elif loop_match_type == 'Covid':
                temporary_collections_counter_on_course_near_match += 1
            success = True
    elif m != xm:
        if (len(el_list[el_list['MMS Id'] == xm]) > 0):
            if ((str(t) == str(xt) or t == xtd) and (xa == a or xa in a or a in xa) and str(yr) == str(xyr)):
                loop_match_type = ave_loop(record)
                if loop_match_type == "non-Covid":
                    ebook_match_on_list_counter += 1
                elif loop_match_type == 'Covid':
                    temporary_collections_portfolio_counter_on_course += 1
                success = True
            elif ((str(t) == str(xt) or t == xtd) and (xa == a or xa in a or a in xa)):
                loop_match_type = ave_loop(record)
                if loop_match_type == "non-Covid":
                    ebook_match_on_list_counter_without_year += 1
                elif loop_match_type == 'Covid':
                    temporary_collections_counter_on_course_near_match += 1
                success = True
        else:

            ebooks_to_add, ebooks_to_add_different_year, ebooks_we_need, covid_e_books_df, covid_e_books_near_match_df
            if ((str(t) == str(xt) or t == xtd) and (xa == a or xa in a or a in xa) and str(yr) == str(xyr)):
                loop_match_type = ave_loop(record)
                if loop_match_type == "non-Covid":
                    ebook_for_physical_counter += 1
                    base_series = course_df.iloc[y]
                    add_series = pd.Series({'Match MMS ID': xml_mms_id, 'Match Title': xml_title, 'Match Author': xml_author, 'Match Publication Year': xml_year, 'Match URL or Collection': url})
                    series_to_add = base_series.append(add_series)
                    ebooks_to_add = ebooks_to_add.append(series_to_add, ignore_index=True)
                elif loop_match_type == 'Covid':
                    base_series = course_df.iloc[y]
                    add_series = pd.Series({'Match MMS ID': xml_mms_id, 'Match Title': xml_title, 'Match Author': xml_author, 'Match Publication Year': xml_year, 'Match URL or Collection': url})
                    series_to_add = base_series.append(add_series)
                    #series_to_add = base_series.append({'Match MMS ID': xml_mms_id, 'Match Title': xml_title, 'Match Author': xml_author, 'Match Publication Year': xml_year, 'Match URL or Collection': url})
                    covid_e_books_df = covid_e_books_df.append(series_to_add, ignore_index=True)
                    #ebooks_to_add = ebooks_to_add.append({'MMS ID': xml_mms_id, 'Title': xml_title, 'Author': xml_author, 'Publication Year': xml_year, 'URL or Collection': xml_url + "|" + collection}, ignore_index=True)
                    temporary_collections_portfolio_counter += 1
                success = True
            elif ((str(t) == str(xt) or t == xtd) and (xa == a or xa in a or a in xa)):
                loop_match_type = ave_loop(record)
                if loop_match_type == "non-Covid":
                    base_series = course_df.iloc[y]
                    add_series = pd.Series({'Match MMS ID': xml_mms_id, 'Match Title': xml_title, 'Match Author': xml_author, 'Match Publication Year': xml_year, 'Match URL or Collection': url})
                    series_to_add = base_series.append(add_series)
                    #series_to_add = base_series.append( Author': xml_author, 'Match Publication Year': xml_year, 'Match URL or Collection': url})
                    ebooks_to_add_different_year = ebooks_to_add_different_year.append(series_to_add, ignore_index=True)
                    #ebooks_to_add_different_year = ebooks_to_add_different_year.append({'Course Code': course_code, 'Course Name': course_name, 'MMS ID': xml_mms_id, 'Title': xml_title, 'Author': xml_author, 'Publication Year': xml_year, 'URL or Collection': url}, ignore_index=True)
                    different_year_ebook_for_physical_counter += 1
                elif loop_match_type == 'Covid':
                    base_series = course_df.iloc[y]
                    add_series = pd.Series({'Match MMS ID': xml_mms_id, 'Match Title': xml_title, 'Match Author': xml_author, 'Match Publication Year': xml_year, 'Match URL or Collection': url})
                    series_to_add = base_series.append(add_series)
                    temporary_collections_portfolio_counter_near_match += 1

                    covid_e_books_near_match_df = covid_e_books_near_match_df.append(series_to_add, ignore_index=True)
                success = True
    return([success, [ebook_match_on_list_counter, ebook_match_on_list_counter_without_year, temporary_collections_portfolio_counter_on_course, temporary_collections_counter_on_course_near_match, ebook_for_physical_counter, different_year_ebook_for_physical_counter, temporary_collections_portfolio_counter, temporary_collections_portfolio_counter_near_match], [ebooks_to_add, ebooks_to_add_different_year, ebooks_we_need, covid_e_books_df, covid_e_books_near_match_df]])


oDir = "./Output"
if not os.path.isdir(oDir) or not os.path.exists(oDir):
    os.makedirs(oDir)



# electronic_filename = askopenfilename(title = "Select electronic inventory comparison filename")
# covid_filename = askopenfilename(title = "Select physical titles with temporary COVID electronic portfolios filename")



# semester = input("Enter the course code prefix for the semester you'd like analyze: ")



chosen_proc_dept_integer = input("\n\nWhat library's reserve lists do you want to analyze.  (blank for all)?\n\n      1) Ginn\n      2) Hirsh\n      3) Music\n      4) SMFA\n      5) Tisch\n      6) Vet\n\nSelect an option: ")

chosen_proc_dept_integer = str(chosen_proc_dept_integer)
chosen_proc_dept = ""
if chosen_proc_dept_integer == "1":
    chosen_proc_dept = "Ginn Reserves"
elif chosen_proc_dept_integer == "2":
    chosen_proc_dept = "Hirsh Health Sciences Reserves"
elif chosen_proc_dept_integer == "3":
    chosen_proc_dept = "Music Reserves"
elif chosen_proc_dept_integer == "4":
    chosen_proc_dept = "SMFA Reserves"
elif chosen_proc_dept_integer == "5":
    chosen_proc_dept = "Tisch Reserves"
elif chosen_proc_dept_integer == "6":
    chosen_proc_dept = "Vet Reserves"
elif chosen_proc_dept_integer == "":
    chosen_proc_dept == ""
else:
    print("You made an invalid selection")
    exit(1)


# Ginn Reserves
# Hirsh Health Sciences Reserves
# Hirsh Reserves
# Music Reserves
# SMFA Reserves
# Tisch Reserves
# Vet Reserves
# semester = str(semester)

reserves_filename = askopenfilename(title = "Select reserves filename")

start = datetime.datetime.now()
#reserves_df = pd.read_csv(reserves_filename, quotechar='"', dtype={'MMS Id': 'str', 'Title (Complete)': 'str', 'Barcode': 'str'}, delimiter=',')

reserves_df = pd.read_excel(reserves_filename, dtype={'MMS Id': 'str', 'Publication Date': 'str', 'Title (Normalized)': 'str'})

reserves_df_total = reserves_df.copy()
reserves_df_total = reserves_df_total.dropna(subset=['MMS Id'])
# reserves_df_total = reserves_df_total.dropna()

reserves_df_total['Title (Normalized)'] = reserves_df_total['Title (Normalized)'].apply(lambda x: x.lower())

# print(reserves_df_total)

pd.set_option('display.max_columns', None)

# print(reserves_df_total)
# reserves_df_total['Semester'] = reserves_df_total['Course Code'].apply(lambda x: re.sub(r'^(\d{4}).+', r'\1', x))
# reserves_df_selected = reserves_df_total[reserves_df_total['Semester'] == semester]
reserves_df_selected = reserves_df_total.copy()
reserves_df_selected = reserves_df_selected.fillna("")
for column in reserves_df_selected.columns:
    reserves_df_selected[column] = reserves_df_selected[column].apply(lambda x: str(x))
    print(column)
    reserves_df_selected[column] = reserves_df_selected[column].apply(lambda x: re.sub(r'[\(\)]', '', x))

reserves_df_selected = reserves_df_selected.sort_values(by=['Processing Department', 'Course Code'])
# print(reserves_df_selected)
column_list = reserves_df_selected.columns
proc_dept_df_list = []
proc_dept_list = []

if chosen_proc_dept == "Hirsh Health Sciences Reserves":
    hhsl_proc_dept_df = reserves_df_selected[reserves_df_selected['Processing Department'] == 'Hirsh Health Sciences Reserves']
    hhsl_proc_dept_df.append(reserves_df_selected[reserves_df_selected['Processing Department'] == 'Hirsh Reserves'])
    hhsl_proc_dept = "All Hirsh Reserves"
    proc_dept_df_list.append(hhsl_proc_dept_df)
    proc_dept_list.append(hhsl_proc_dept)

elif chosen_proc_dept != "":
    proc_dept_df_list.append(reserves_df_selected[reserves_df_selected['Processing Department'] == chosen_proc_dept])
    proc_dept_list.append(chosen_proc_dept)

else:
    for proc_dept, proc_dept_df in reserves_df_selected.groupby('Processing Department'):
        proc_dept_df_list.append(proc_dept_df)
        proc_dept_list.append(proc_dept)



d = 0

# print("got to line 68\n")
# test_file = open(oDir + '/Test File.txt', 'w+', encoding='utf-8')
# print(proc_dept_df_list)
# print(str(type(proc_dept_df_list)))
# print(str(d))
# print(proc_dept_df_list[d])
# sys.exit()
for dept in proc_dept_df_list:

    processing_dept = proc_dept_list[d]
    if processing_dept == 'Hirsh Reserves' or processing_dept == 'Hirsh Health Sciences Reserves':
        processing_dept = "All Hirsh Reserves"
    date = datetime.datetime.now().strftime("%m-%d-%Y")
    filename = oDir + '/Electronic Copy Counts for - ' + processing_dept + ' - ' + date + '.xlsx'


    wb = openpyxl.Workbook()
    wb.save(filename)
    workbook = load_workbook(filename)
    writer = pd.ExcelWriter(filename, engine='openpyxl')
    writer.book = workbook
    # workbook = writer.book
    # counts_sheet = workbook.create_sheet('Counts')
    # counts_sheet = workbook.create_sheet('InRepo')
    # counts_sheet = workbook.create_sheet('InRepoDifferentYear')
    # counts_sheet = workbook.create_sheet('TempColl')
    # counts_sheet = workbook.create_sheet('TempCollDiffYear')
    # counts_sheet = workbook.create_sheet('ForPurchase')

    #
    #
    #
    # in_repo = workbook.add_worksheet('InRepo')
    # in_repo_different_year = workbook.add_worksheet('InRepoDifferentYear')
    #
    # temporary_collection_sheet = workbook.add_worksheet('TempColl')
    # temporary_collection_diff_year_sheet = workbook.add_worksheet('TempCollDiffYear')
    #
    # to_order_sheet = workbook.add_worksheet('ForPurchase')

    if d >= 1:
        break
    course_df_list = []
    # reserves_df_selected = reserves_df_selected[reserves_df_selected['Processing Department'] == 'Hirsh Health Sciences Reserves']
    for course, course_df in proc_dept_df_list[d].groupby('Course Code'):
        course_df_list.append(course_df)

    ## testing


    x = 0
    col_list = reserves_df_selected.columns.tolist()
    output_cols = reserves_df_selected.columns.tolist()

    output_cols.extend(['Match MMS ID', 'Match Title', 'Match Author', 'Match Publication Year', 'Match URL or Collection'])
    counts_df = pd.DataFrame(columns=['Processing Department', 'Books on Course', 'Physical Books on Course', 'No Electronic Version for Physical Book', 'Electronic - Already on Course', 'Electronic - Already on Course - Different Year', 'Electronic - Already on Course - COVID Temporary Electronic Collection', 'Electronic - Already on Course - COVID Temporary Electronic Collection - Different Year', 'Electronic - In Collection - Add to Course', 'Electronic - In Collection - Potentially Add to Course - Different Year', 'Electronic - Temporarily in Collection', 'Electronic - Temporarily in Collection - Different Year'])
    ebooks_to_add = pd.DataFrame(columns=output_cols)
    ebooks_to_add_different_year = pd.DataFrame(columns=output_cols)
    # ebooks_on_course = pd.DataFrame(columns=output_cols)
    # ebooks_on_course_different_year = pd.DataFrame(columns=output_cols)

    ebooks_we_need = pd.DataFrame(columns=col_list)
    #temporary_physical_record_ebooks = pd.DataFrame(columns=column_list)

    sru_url_prefix = "https://tufts.alma.exlibrisgroup.com/view/sru/01TUN_INST?version=1.2&operation=searchRetrieve&recordSchema=marcxml&query="
    # covid_column_list = column_list
    # covid_column_list = covid_column_list.append(['URL or Collection'])

    covid_e_books_df = pd.DataFrame(columns=output_cols)

    covid_e_books_near_match_df = pd.DataFrame(columns=output_cols)
    while x < len(course_df_list):
        # print(course_df_list[x])
        # print("Got into course_df_list loop")
        # print(course_df_list)
        # print("\n\n\n\n\n")
        # print(course_df_list[x])
        course_name = course_df_list[x]['Course Name']

        # if x >= 5:
        #     break
        y = 0
        temporary_collections_portfolio_counter = 0
        temporary_collections_portfolio_counter_on_course = 0
        temporary_collections_counter_on_course_near_match = 0
        temporary_collections_portfolio_counter_near_match = 0
        ebook_counter = 0
        ebook_for_physical_counter = 0
        different_year_ebook_for_physical_counter = 0
        ebook_match_on_list_counter_without_year = 0

        ebook_match_on_list_counter = 0
        ebooke_match_on_list_counter_without_year = 0
        no_match_counter = 0
        number_of_books_on_course = len(course_df_list[x])


        date = datetime.datetime.now().strftime("%m/%d/%Y")

        #counts_file = open(oDir + '/Counts - ' + str(course_name) + ' - ' + date + '.txt', 'w+')
        counts_per_course_list = []

        # print("Items on course: " + str(len(course_df_list[x])))
        course_ebook_count = 0
        course_code = course_df_list[x].iloc[0]['Course Code']
        course_name = course_df_list[x].iloc[0]['Course Name']

        print(course_code + "\n" + course_name + "\n")


        physical_list = course_df_list[x][course_df_list[x]['Resource Type'] == 'Book - Physical']


        electronic_list = course_df_list[x][course_df_list[x]['Resource Type'] == 'Book - Electronic']

        number_of_physical_books_on_course = len(physical_list)

        # print(physical_list)
        if len(physical_list) > 0:

            while y < len(physical_list):
                match = False
                course_df = physical_list.copy()
                # print(course_df)
                # print(course_df)
                # if pd.isnull(course_df['Title (Normalized)']):
                #     print("Null: " + course_df[y]['Title (Normalized)'])
                # print(physical_list)
                # if x > 20:
                #     break

                # test_file.write(str(course_df.iloc[y]) + "\n")

                # if y > 2:
                #     break
                # print("Resource Type: " + str(course_df.iloc[y]['Resource Type']))


                title = course_df.iloc[y]['Title (Normalized)']
                title_for_query = '\"' + re.sub(r'\s', '%20', title) + '\"'
                author = ""
                # print(str(author) + "Type: " + str(type(author)))
                if course_df.iloc[y]['Author'] != "" and course_df.iloc[y]['Author'] is not np.nan:
                    #author = re.sub(r'([^,;]+,\s+[^,;],*\s*[^;,]*).+', r'\1', course_df.iloc[y]['Author'])
                    author = re.sub(r'(,\sauthor\.?|,\scontributor\.?|\scontributor\.?|\sauthor\.?)', '', course_df.iloc[y]['Author'])
                    author = author.lower()
                elif course_df.iloc[y]['Author (contributor)'] != "" and course_df.iloc[y]['Author (contributor)'] is not np.nan:
                    author = re.sub(r'(,\sauthor|,\scontributor|\scontributor|\sauthor)', '', course_df.iloc[y]['Author (contributor)'])
                    author = author.lower()
                #author_for_query = '\"' + re.sub(r'\s', '%20', author) + '\"'
                year = course_df.iloc[y]['Publication Date']
                year = re.sub(r'\D', '', year)
                mms_id = course_df.iloc[y]['MMS Id']
                # if course_df.iloc[y]['Resource Type'] == 'Book - Electronic':
                #     ebook_counter += 1
                #     ebooks_on_this_course.loc[len(ebooks_on_this_course)] = course_df.iloc[y]
                #
                #     course_ebook_count += 1
                #     continue
                # elif course_df.iloc[y]['Resource Type'] == 'Book - Physical':
                #     if len(course_df[(course_df['Title (Normalized)'] == title) and ([course_df['Author'] = author) and (course_df['Publication Year'] == 'year')] > 1:
                #         ebook_for_physical_counter += 1
                #         continue
                #     elif len(course_df[(course_df['Title (Normalized)'] == title) and ([course_df['Author'] = author)] > 1:
                #         different_year_ebook_for_physical_counter += 1
                # author_list_string = '; '.join(electronic_list['Author'].tolist())
                # contributor_list_string = '; '.join(electronic_list['Author'].tolist())


                if len(electronic_list) >= 1:
                    # if course_code == "2202-20870":
                    #     print(electronic_list)

                    electronic_list['Author'] = electronic_list['Author'].apply(lambda x: re.sub(r'(,\sauthor|,\scontributor|\scontributor|\sauthor)', '', x).lower())
                    electronic_list['Author (contributor)'] = electronic_list['Author (contributor)'].apply(lambda x: re.sub(r'(,\sauthor|,\scontributor|\scontributor|\sauthor)', '', x).lower())
                # if course_code == "2202-20859":
                #     print("################\n################\n################\n################\n################\n")
                #
                #     print(electronic_list['Author'])
                #     print(electronic_list['Author (contributor)'])
                #     print("################\n")
                #     print(author)
                #     print("################\n")

                f = 0
                for f in range(0, len(electronic_list)):

                    if (electronic_list.iloc[f]['Title (Normalized)'] == title) and (electronic_list.iloc[f]['Author (contributor)'] in author or author in electronic_list.iloc[f]['Author (contributor)'] or electronic_list.iloc[f]['Author'] in author or author in electronic_list.iloc[f]['Author']) and (electronic_list.iloc[f]['Publication Date'] == year):
                        ebook_match_on_list_counter += 1
                        match = True
                        y += 1
                        break

                    elif (electronic_list.iloc[f]['Title (Normalized)'] == title) and (electronic_list.iloc[f]['Author (contributor)'] in author or author in electronic_list.iloc[f]['Author (contributor)'] or electronic_list.iloc[f]['Author'] in author or author in electronic_list.iloc[f]['Author']):
                        ebook_match_on_list_counter_without_year += 1

                        quasi_match_bool = True
                        match = True
                        y += 1
                        break
                    f += 1
                if match == True:
                    continue
                else:
                # electronic_match = electronic_list[(electronic_list['Title (Normalized)'] == title) & (electronic_list['Author (contributor)'].str.contains(author) | electronic_list['Author'].str.contains(author)) & (electronic_list['Publication Date'] == year)]
                # electronic_match_without_year = electronic_list[(electronic_list['Title (Normalized)'] == title) & (electronic_list['Author (contributor)'].str.contains(author) | electronic_list['Author'].str.contains(author))]
                # quasi_match_bool = False
                #
                # if len(electronic_match) >= 1:
                #     ebook_match_on_list_counter += 1
                #     match = True
                #     y += 1
                #     continue
                #
                # elif len(electronic_match_without_year) >= 1:
                #




                    record_result = requests.get(sru_url_prefix + 'alma.title=' + title_for_query)
                    #print(record_result.content)
                    tree = et.ElementTree(et.fromstring(record_result.content))
                    root = tree.getroot()
                    # print(tree)
                    # print(root)
                    z = 0
                    # result_df = pd.DataFrame(columns=['MMS Id', 'Title', 'Author', 'Year'])

                    for elem in root.iter():


                        if re.match(r'.*record$', elem.tag):
                            # print(elem.tag)
                            # print(et.tostring(elem))
                            bib_record = pym.parse_xml_to_array(io.StringIO(et.tostring(elem).decode('utf-8')))
                            # print(bib_record[0])
                            # test_file.write(str(bib_record[0]) + "\n")
                            # print(course_df.iloc[y])
                            # print(bib_record[0])
                            if '590' in bib_record[0]:
                                if 'a' in bib_record[0]['590']:
                                    if bib_record[0]['590']['a'].lower() == 'on the fly':
                                        continue
                            if '001' in bib_record[0]:
                                # print("got into if statement")

                                # if bib_record[0]['001'] == mms_id:
                                #     z += 1
                                #     continue


                                # if '655' not in bib_record[0]:
                                #     z += 1
                                #     continue
                                # elif 'Electronic books' not in bib_record[0]['655']['a']:
                                #     continue
                                xml_mms_id = ""

                                # print("Data: " + str(bib_record[0]['001']['data']))
                                # print("Data value: " + bib_record[0]['001'].value())
                                # for t in bib_record[0]['001']:
                                #     print("T: " + str(t))
                                # if 'data' in bib_record[0]['001']:
                                #     for s in bib_record[0]['001']['data']:
                                #         print("S: " + str(s))

                                xml_mms_id = bib_record[0]['001'].value()
                                # print('Analytics ID: ' + str(mms_id))
                                # print('XML MMS ID:   ' + str(xml_mms_id))


                                xml_url = ""

                                if '856' in bib_record[0]:
                                    if 'u' in bib_record[0]['856']:
                                        xml_url = bib_record[0]['856']['u']
                                a = ""
                                b = ""
                                if 'a' in bib_record[0]['245']:
                                    a = bib_record[0]['245']['a']
                                if 'b' in bib_record[0]['245']:
                                    b = bib_record[0]['245']['b']

                                xml_title = a + b
                                xml_title = xml_title.lower()
                                xml_title = re.sub(r'\s\/$', '', xml_title)
                                xml_title = re.sub(r'\'', ' ', xml_title)

                                xml_title = re.sub(r'\.$', '', xml_title)
                                xml_title = re.sub(r'^(the\s|a\s)', '', xml_title)
                                xml_title = re.sub(r'\s{2,}', ' ', xml_title)
                                xml_title_dash = xml_title
                                xml_title = re.sub(r'[^a-zA-Z0-9 ]', ' ', xml_title)
                                xml_title_dash = re.sub(r'[^a-zA-Z0-9 -_]', ' ', xml_title)
                                xml_title = re.sub(r'\s{2,}', ' ', xml_title)
                                xml_title_dash = re.sub(r'\s{2,}', ' ', xml_title_dash)

                                xml_author = ""

                                if '100' in bib_record[0]:
                                    if 'a' in bib_record[0]['100']:
                                        xml_author = bib_record[0]['100']['a']
                                    if 'd' in bib_record[0]['100']:
                                        xml_author += " " + bib_record[0]['100']['d']
                                    if 'e' in bib_record[0]['100']:
                                        xml_author += " " + bib_record[0]['100']['e']
                                elif '110' in bib_record[0]:
                                    if 'a' in bib_record[0]['110']:
                                        xml_author = bib_record[0]['110']['a']
                                    if 'd' in bib_record[0]['110']:
                                        xml_author += " " + bib_record[0]['110']['d']
                                    if 'e' in bib_record[0]['110']:
                                        xml_author += " " + bib_record[0]['110']['e']
                                elif '111' in bib_record[0]:
                                    if 'a' in bib_record[0]['111']:
                                        xml_author = bib_record[0]['111']['a']
                                    if 'd' in bib_record[0]['111']:
                                        xml_author += " " + bib_record[0]['111']['d']
                                    if 'e' in bib_record[0]['111']:
                                        xml_author += " " + bib_record[0]['111']['e']
                                elif '700' in bib_record[0]:
                                    if 'a' in bib_record[0]['700']:
                                        xml_author = bib_record[0]['700']['a']
                                    if 'd' in bib_record[0]['700']:
                                        xml_author += " " + bib_record[0]['700']['d']
                                    if 'e' in bib_record[0]['700']:
                                        xml_author += " " + bib_record[0]['700']['e']
                                elif '710' in bib_record[0]:
                                    if 'a' in bib_record[0]['710']:
                                        xml_author = bib_record[0]['710']['a']
                                    if 'd' in bib_record[0]['710']:
                                        xml_author += " " + bib_record[0]['710']['d']
                                    if 'e' in bib_record[0]['710']:
                                        xml_author += " " + bib_record[0]['710']['e']
                                elif '711' in bib_record[0]:
                                    if 'a' in bib_record[0]['711']:
                                        xml_author = bib_record[0]['711']['a']
                                    if 'd' in bib_record[0]['711']:
                                        xml_author += " " + bib_record[0]['711']['d']
                                    if 'e' in bib_record[0]['711']:
                                        xml_author += " " + bib_record[0]['711']['e']

                                # xml_author = xml_author.lower()
                                # xml_author = re.sub(r',$', '', xml_author)
                                xml_author = re.sub(r'(,\sauthor\.?|,\scontributor\.?|\scontributor\.?|\sauthor\.?)', '', xml_author)
                                xml_author = xml_author.lower()

                                xml_year = ""

                                if '260' in bib_record[0]:
                                    if 'c' in bib_record[0]['260']:
                                        xml_year = bib_record[0]['260']['c']
                                        xml_year = re.sub(r'\D', '', xml_year)
                                elif '264' in bib_record[0]:
                                    if 'c' in bib_record[0]['264']:
                                        xml_year = bib_record[0]['264']['c']
                                        xml_year = re.sub(r'\D', '', xml_year)

                                collection = ""
                                type = ""
                                if 'AVE' in bib_record[0]:
                                    if 'm' in bib_record[0]['AVE']:
                                        collection = bib_record[0]['AVE']['m']
                                        type = 'temp'
                                url = ""
                                if collection != "" and xml_url != "":
                                    url =  xml_url + "|" + collection
                                elif collection != "":
                                    url = collection
                                elif xml_url != "":
                                    url = xml_url

                                electronic_h_bool = False
                                if 'h' in bib_record[0]['245']:
                                    if 'electronic resource' in bib_record[0]['245']['h'] or 'Electronic resource' in bib_record[0]['245']['h']:
                                        electronic_h_bool = True
                                        type = 'electronic_245_h'


                                if '655' in bib_record[0] and "Electronic books".lower() in bib_record[0]['655']['a'].lower():
                                    type = 'electronic_655'


                                # if '655' in bib_record[0] and "Electronic books".lower() in bib_record[0]['655']['a'].lower():
                                #     type = 'ebook'
                                # print("\n\n\nXML:        " + str(xml_title) + "|" + str(xml_author) + "|" + str(xml_year))
                                # print(      "Analytics:  " + str(title) + "|" + str(author) + "|" + str(year) + "|" + type + "\n\n\n")
                                # test_file.write("\n\n\nXML:        " + str(xml_title) + "|" + str(xml_author) + "|" + str(xml_year) + "\n")
                                # test_file.write(      "Analytics:  " + str(title) + "|" + str(author) + "|" + str(year) + "|" + type + "\n\n\n")
                                if 'AVE' in bib_record[0]:
                                    return_list = ebook_match(bib_record[0], mms_id, title, author, year, xml_mms_id, xml_title, xml_title_dash, xml_author, xml_year, electronic_list, ebook_match_on_list_counter, ebook_match_on_list_counter_without_year, temporary_collections_portfolio_counter_on_course, temporary_collections_counter_on_course_near_match, ebook_for_physical_counter, different_year_ebook_for_physical_counter, temporary_collections_portfolio_counter, temporary_collections_portfolio_counter_near_match, y, course_df, ebooks_to_add, ebooks_to_add_different_year, ebooks_we_need, covid_e_books_df, covid_e_books_near_match_df)
                                    match = return_list[0]
                                    counts_return = return_list[1]
                                    dataframe_return = return_list[2]

                                    #counts
                                    ebook_match_on_list_counter = counts_return[0]
                                    ebook_match_on_list_counter_without_year = counts_return[1]
                                    temporary_collections_portfolio_counter_on_course = counts_return[2]
                                    temporary_collections_counter_on_course_near_match = counts_return[3]
                                    ebook_for_physical_counter = counts_return[4]
                                    different_year_ebook_for_physical_counter = counts_return[5]
                                    temporary_collections_portfolio_counter = counts_return[6]
                                    temporary_collections_portfolio_counter_near_match = counts_return[7]

                                    #dataframes
                                    ebooks_to_add = dataframe_return[0]
                                    ebooks_to_add_different_year = dataframe_return[1]
                                    ebooks_we_need = dataframe_return[2]
                                    covid_e_books_df = dataframe_return[3]
                                    covid_e_books_near_match_df = dataframe_return[4]


                                # if str(mms_id) == str(xml_mms_id):
                                #     if ("AVE" in bib_record[0] and (str(title) == str(xml_title) or title == xml_title_dash) and (xml_author == author or xml_author in author or author in xml_author) and str(year) == str(xml_year)):
                                #         perm_bool = False
                                #         for ave in bib_record[0].get_fields('AVE'):
                                #             if ('Covid-19'.lower() not in ave['m'].lower()):
                                #                 perm_bool = True
                                #
                                #                 ebook_match_on_list_counter += 1
                                #                 # temporary_collections_portfolio_counter_on_course += 1
                                #                 match = True
                                #                 break
                                #         if perm_bool == False:
                                #             temporary_collections_portfolio_counter_on_course += 1
                                #             match = True
                                #         break
                                # else:
                                # elif ("AVE" in bib_record[0] and (str(title) == str(xml_title) or title == xml_title_dash) and (xml_author == author or xml_author in author or author in xml_author) and (str(year) == str(xml_year)) and (str(mms_id) != str(xml_mms_id))):
                                #
                                #
                                #         if ('Covid-19'.lower() not in ave['m'].lower()):
                                #             perm_bool = True
                                #             base_series = course_df.iloc[y]
                                #             add_series = pd.Series({'Match MMS ID': xml_mms_id, 'Match Title': xml_title, 'Match Author': xml_author, 'Match Publication Year': xml_year, 'Match URL or Collection': url})
                                #             series_to_add = base_series.append(add_series)
                                #             ebooks_to_add = ebooks_to_add.append(series_to_add, ignore_index=True)
                                #             ebook_for_physical_counter += 1
                                #             match = True
                                #             break
                                #     if perm_bool == False:
                                #
                                #
                                #         if (('AVE' in bib_record[0] and 'Covid-19'.lower() in collection.lower()) and (str(title) == str(xml_title) or title == xml_title_dash) and (xml_author == author or xml_author in author or author in xml_author) and str(year) == str(xml_year) and str(mms_id) == str(xml_mms_id)):
                                #             print('Got into temp colleciton match on course')
                                #             ebook_match_on_list_counter += 1
                                #             temporary_collections_portfolio_counter_on_course += 1
                                #             match = True
                                #             break
                                #         elif (('AVE' in bib_record[0] and 'Covid-19'.lower() in collection.lower()) and (str(title) == str(xml_title) or title == xml_title_dash) and (xml_author == author or xml_author in author or author in xml_author) and str(mms_id) == str(xml_mms_id)):
                                #             #('AVE' in bib_record[0] and str(title) == str(xml_title) and (xml_author == author or xml_author in author or author in xml_author) and str(year) == str(xml_year)):
                                #             print('Got into temp colleciton match on course without year')
                                #             ebook_match_on_list_counter_without_year += 1
                                #             temporary_collections_counter_on_course_near_match += 1
                                #             match = True
                                #             break
                                #         elif (('AVE' in bib_record[0] and 'Covid-19'.lower() in collection.lower()) and (str(title) == str(xml_title) or title == xml_title_dash) and (xml_author == author or xml_author in author or author in xml_author) and str(year) == str(xml_year)):
                                #             # print("got into AVE")
                                #             print('Got into temp colleciton match in repo')
                                #             base_series = course_df.iloc[y]
                                #             add_series = pd.Series({'Match MMS ID': xml_mms_id, 'Match Title': xml_title, 'Match Author': xml_author, 'Match Publication Year': xml_year, 'Match URL or Collection': url})
                                #             series_to_add = base_series.append(add_series)
                                #             #series_to_add = base_series.append({'Match MMS ID': xml_mms_id, 'Match Title': xml_title, 'Match Author': xml_author, 'Match Publication Year': xml_year, 'Match URL or Collection': url})
                                #             ebooks_to_add = ebooks_to_add.append(series_to_add, ignore_index=True)
                                #             #ebooks_to_add = ebooks_to_add.append({'MMS ID': xml_mms_id, 'Title': xml_title, 'Author': xml_author, 'Publication Year': xml_year, 'URL or Collection': xml_url + "|" + collection}, ignore_index=True)
                                #             temporary_collections_portfolio_counter += 1
                                #             ebook_for_physical_counter += 1
                                #             #ebook_match_on_list_counter += 1
                                #             #
                                #             #series_to_add = base_series.append({'Match MMS ID': xml_mms_id, 'Match Title': xml_title, 'Match Author': xml_author, 'Match Publication Year': xml_year, 'Match URL or Collection': url})
                                #             covid_e_books_df = covid_e_books_df.append(series_to_add, ignore_index=True)
                                #             #covid_e_books_df = covid_e_books_df.append({'Course Code': course_code, 'Course Name': course_name, 'MMS ID': xml_mms_id, 'Title': xml_title, 'Author': xml_author, 'Publication Year': xml_year, 'URL or Collection':  url}, ignore_index=True)
                                #
                                #             match = True
                                #             break
                                #
                                #         elif (('AVE' in bib_record[0] and 'Covid-19'.lower() in collection.lower()) and (str(title) == str(xml_title) or title == xml_title_dash) and (xml_author == author or xml_author in author or author in xml_author)):
                                #             #print('Got into temp colleciton match in repo without year')
                                #             base_series = course_df.iloc[y]
                                #             add_series = pd.Series({'Match MMS ID': xml_mms_id, 'Match Title': xml_title, 'Match Author': xml_author, 'Match Publication Year': xml_year, 'Match URL or Collection': url})
                                #             series_to_add = base_series.append(add_series)
                                #             temporary_collections_portfolio_counter_near_match += 1
                                #             #series_to_add = base_series.append({'Match MMS ID': xml_mms_id, 'Match Title': xml_title, 'Match Author': xml_author, 'Match Publication Year': xml_year, 'Match URL or Collection': url})
                                #             ebooks_to_add_different_year = ebooks_to_add_different_year.append(series_to_add, ignore_index=True)
                                #             #ebooks_to_add_different_year = ebooks_to_add_different_year.append({'Course Code': course_code, 'Course Name': course_name, 'MMS ID': xml_mms_id, 'Title': xml_title, 'Author': xml_author, 'Publication Year': xml_year, 'URL or Collection':  url}, ignore_index=True)
                                #             #series_to_add = base_series.append({'Match MMS ID': xml_mms_id, 'Match Title': xml_title, 'Match Author': xml_author, 'Match Publication Year': xml_year, 'Match URL or Collection': url})
                                #             covid_e_books_near_match_df = covid_e_books_near_match_df.append(series_to_add, ignore_index=True)
                                #             #covid_e_books_near_match_df = covid_e_books_near_match_df.append({'Course Code': course_code, 'Course Name': course_name, 'MMS ID': xml_mms_id, 'Title': xml_title, 'Author': xml_author, 'Publication Year': xml_year, 'URL or Collection': collection}, ignore_index=True)
                                #             different_year_ebook_for_physical_counter += 1
                                #
                                #             match=True
                                #             break
                                # elif (('655' in bib_record[0] and "Electronic books".lower() in bib_record[0]['655']['a'].lower() or electronic_h_bool == True) and (str(title) == str(xml_title) or title == xml_title_dash) and (xml_author == author or xml_author in author or author in xml_author) and str(year) == str(xml_year)):
                                #     base_series = course_df.iloc[y]
                                #     add_series = pd.Series({'Match MMS ID': xml_mms_id, 'Match Title': xml_title, 'Match Author': xml_author, 'Match Publication Year': xml_year, 'Match URL or Collection': url})
                                #     series_to_add = base_series.append(add_series)
                                #     ebooks_to_add = ebooks_to_add.append(series_to_add, ignore_index=True)
                                #     ebook_for_physical_counter += 1
                                #     match = True
                                #     break
                                # elif (('655' in bib_record[0] and "Electronic books".lower() in bib_record[0]['655']['a'].lower() or electronic_h_bool == True) and (str(title) == str(xml_title) or title == xml_title_dash) and (xml_author in author or author in xml_author) and quasi_match_bool == False):
                                #     base_series = course_df.iloc[y]
                                #     add_series = pd.Series({'Match MMS ID': xml_mms_id, 'Match Title': xml_title, 'Match Author': xml_author, 'Match Publication Year': xml_year, 'Match URL or Collection': url})
                                #     series_to_add = base_series.append(add_series)
                                #     #series_to_add = base_series.append( Author': xml_author, 'Match Publication Year': xml_year, 'Match URL or Collection': url})
                                #     ebooks_to_add_different_year = ebooks_to_add_different_year.append(series_to_add, ignore_index=True)
                                #     #ebooks_to_add_different_year = ebooks_to_add_different_year.append({'Course Code': course_code, 'Course Name': course_name, 'MMS ID': xml_mms_id, 'Title': xml_title, 'Author': xml_author, 'Publication Year': xml_year, 'URL or Collection': url}, ignore_index=True)
                                #     different_year_ebook_for_physical_counter += 1
                                #     match = True
                                #     break


                                z += 1
                    if (xml_mms_id != mms_id and type != ""):
                        print("Match after of XML loop:" + str(match))
                if (xml_mms_id != mms_id and type != ""):
                    print("Match after else statement: " + str(match))
                    print("XML:        " + str(xml_mms_id) + "|" + str(xml_title) + "|" + str(xml_author) + "|" + str(xml_year) + "|" + type)
                    print("Analytics:  " + str(mms_id) + "|" + str(title) + "|" + str(author) + "|" + str(year))
                if not match:
                    print("Defaulted into not found")
                    # base_series = course_df.iloc[y]
                    #add_series = pd.Series({'Course Code': course_code, 'Course Name': course_name, 'MMS ID': mms_id, 'Title': title, 'Author': author, 'Publication Year': year, 'Match MMS ID': xml_mms_id, 'Match Title': xml_title, 'Match Author': xml_author, 'Match Publication Year': xml_year, 'Match URL or Collection': url})
                    # series_to_add = base_series.append(add_series)
                    # series_to_add = base_series.append({'Match MMS ID': xml_mms_id, 'Match Title': xml_title, 'Match Author': xml_author, 'Match Publication Year': xml_year, 'Match URL or Collection': url})
                    ebooks_we_need = ebooks_we_need.append(course_df.iloc[y], ignore_index=True)
                    #ebooks_we_need = ebooks_we_need.append(course_df.iloc[y], ignore_index=True)
                    no_match_counter += 1

                #print(ebooks_on_this_course)
                y += 1
        # print("Course ebook count:               " + str(course_ebook_count))
        # print("Different:                        " + str(different_year_ebook_for_physical_counter))
        # print("Temporary:                        " + str(temporary_collections_portfolio_counter_near_match))
        # print("Ebook for physical:               " + str(ebook_for_physical_counter))
        #

        # print("Number of books on course:                                                " + str(number_of_books_on_course))
        # print("Number of books for which there's an electronic copy on reading list:     " + str(ebook_match_on_list counter))
        # print("     Ebooks on list under temporary COVID Collection")
        #
        course_name = course_df_list[x].iloc[0]['Course Name']
        course_code = course_df_list[x].iloc[0]['Course Code']
        processing_dept = course_df_list[x].iloc[0]['Processing Department']


        #counts_file = open(oDir + '/Counts - ' + str(course_name) + ' - ' + date + '.txt', 'w+')

        counts_df = counts_df.append({'Processing Department': processing_dept, 'Course Name': course_name, 'Course Code': course_code, 'Books on Course': number_of_books_on_course,  'Physical Books on Course' :number_of_physical_books_on_course,  'No Electronic Version for Physical Book': no_match_counter, 'Electronic - Already on Course': ebook_match_on_list_counter, 'Electronic - Already on Course - Different Year': ebook_match_on_list_counter_without_year, 'Electronic - Already on Course - COVID Temporary Electronic Collection': temporary_collections_portfolio_counter_on_course,  'Electronic - Already on Course - COVID Temporary Electronic Collection - Different Year': temporary_collections_counter_on_course_near_match, 'Electronic - In Collection - Add to Course': ebook_for_physical_counter, 'Electronic - In Collection - Potentially Add to Course - Different Year': different_year_ebook_for_physical_counter, 'Electronic - Temporarily in Collection': temporary_collections_portfolio_counter, 'Electronic - Temporarily in Collection - Different Year': temporary_collections_portfolio_counter_near_match}, ignore_index=True)

        # processing_dept_master = proc_dept_df_list[d].iloc[0]['Processing Department']
        #

        # date = datetime.datetime.now().strftime("%m-%d-%Y")
        # filename = oDir + '/Electronic Copy Counts for - ' + processing_dept + ' - ' + date + '.xlsx'
        #
        # book = load_workbook(filename)
        # writer = pandas.ExcelWriter(filename, engine='xlsxwriter')
        # writer.book = book
        # counts_sheet writer.sheets['Sheet'] = dict((ws.title, ws) for ws in book.worksheets



        # counts_sheet = sheet_1
        #
        # in_repo = workbook.add_worksheet('InRepo')
        # in_repo_different_year = workbook.add_worksheet('InRepoDifferentYear')
        #
        # temporary_collection_sheet = workbook.add_worksheet('TempColl')
        # temporary_collection_sheet = workbook.add_worksheet('TempCollDiffYear')
        #
        # to_order_sheet = workbook.add_worksheet('ForPurchase')



        # counts_df.drop(['Processing Department', 'Ebook Copy for Physical in Repo - Potentially Add - Different Year'])
        # if x == 0:
        #     counts_df.to_excel(writer, sheet_name='Counts', startrow=0, index=False)
        #     ebooks_to_add.to_excel(writer, sheet_name='InRepo',  startrow=0, index=False)
        #     ebooks_to_add_different_year.to_excel(writer, sheet_name='InRepoDifferentYear', startrow=0, index=False)
        #     covid_e_books_df.to_excel(writer, sheet_name='TempColl', startrow=0, index=False)
        #     covid_e_books_near_match_df.to_excel(writer, sheet_name='TempCollDiffYear', startrow=0, index=False)
        #     ebooks_to_add.to_excel(writer, sheet_name='ForPurchase', startrow=0, index=False)
        #     writer.save()
        # else:
        #
        #     counts_max_row = writer.book['Counts'].max_row + 1
        #     in_repo_max_row = writer.book['InRepo'].max_row + 1
        #     in_repo_diff_max_row = writer.book['InRepoDifferentYear'].max_row + 1
        #     temp_coll_max_row = writer.book['TempColl'].max_row + 1
        #     temp_coll_diff_year_max_row = writer.book['InRepoDifferentYear'].max_row + 1
        #     to_order_max_row = writer.book['ForPurchase'].max_row + 1
        #     print(
        #     "Counts - Type: " + str(counts_max_row) + "\n" +
        #     "InRepo: " + str(in_repo_max_row) + "\n" +
        #     "InRepoDifferentYear: " + str(in_repo_diff_max_row) + "\n" +
        #     "TempColl: " + str(temp_coll_max_row) + "\n" +
        #     "InRepoDifferentYear: " + str(temp_coll_max_row) + "\n"
        #     "ForPurchase: " + str(to_order_max_row) + "\n"
        #     )
        #
        #     max_counts_row_by_x = x + 1
        #     counts_df.to_excel(writer, sheet_name='Counts', startrow=x + 1, index=False, header=False, engine='openpyxl')
        #     ebooks_to_add.to_excel(writer, sheet_name='InRepo', startrow=in_repo_max_row, index=False, header=False, engine='openpyxl')
        #     ebooks_to_add_different_year.to_excel(writer, sheet_name='InRepoDifferentYear', startrow=in_repo_diff_max_row, index=False, header=False, engine='openpyxl')
        #     covid_e_books_df.to_excel(writer, sheet_name='TempColl', startrow=temp_coll_max_row, index=False, header=False, engine='openpyxl')
        #     covid_e_books_near_match_df.to_excel(writer, sheet_name='TempCollDiffYear', startrow=temp_coll_diff_year_max_row, index=False, header=False, engine='openpyxl')
        #     ebooks_to_add.to_excel(writer, sheet_name='ForPurchase', startrow=to_order_max_row, index=False, header=False, engine='openpyxl')
        #     writer.save()
        #     book.save(filename)
        x += 1
    # This section moves the last two columns, course data, to the first two positions, and puts proc dept in beginning
    cols = counts_df.columns.tolist()
    cols = cols[-1:] + cols[:-1]
    counts_df = counts_df[cols]
    cols = cols[-1:] + cols[:-1]
    counts_df = counts_df[cols]
    cols.remove('Processing Department')
    cols.insert(0, 'Processing Department')
    counts_df = counts_df[cols]




    counts_df.to_excel(writer, sheet_name='Counts', index=False, engine='openpyxl')
    ebooks_to_add.to_excel(writer, sheet_name='InRepo', index=False, engine='openpyxl')
    ebooks_to_add_different_year.to_excel(writer, sheet_name='InRepoDifferentYear', index=False, engine='openpyxl')
    covid_e_books_df.to_excel(writer, sheet_name='TempColl', index=False, engine='openpyxl')
    covid_e_books_near_match_df.to_excel(writer, sheet_name='TempCollDiffYear', index=False, engine='openpyxl')
    ebooks_we_need.to_excel(writer, sheet_name='NeededNotInTufts', index=False, engine='openpyxl')

    ebooks_we_need_unique = ebooks_we_need.drop_duplicates(subset=['MMS Id'])

    # print("\n\n\n")
    # print(ebooks_we_need_unique)

    ebooks_we_need_unique.to_excel(writer, sheet_name='UniquePrchsList', index=False, engine='openpyxl')




    for sheet in writer.sheets:
        e = 0
        for column in writer.sheets[sheet].iter_cols():
            writer.sheets[sheet].column_dimensions[get_column_letter(e + 1)].width = "20"
            e += 1
    f = 0

    for sheet in writer.sheets:
        writer.sheets[sheet].freeze_panes = 'A2'
        for row in writer.sheets[sheet]:
            # print(row)
            if f == 0:

                for cell in row:
                    #cell.style.alignment.wrap_text=True
                    cell.alignment = Alignment(wrap_text=True)
                    cell.font = Font(bold=True)


            f += 1

    # sums
    counts_max_row = writer.book['Counts'].max_row

    for x in range(1, 13):
        writer.sheets['Counts'].cell(row = counts_max_row + 2, column = x).font = Font(bold=True)

    writer.sheets['Counts'].cell(row = counts_max_row + 2, column = 1).value = "Totals"
    writer.sheets['Counts'].cell(row = counts_max_row + 2, column = 4).value = '= SUM(D1:D' + str(counts_max_row) + ')'
    writer.sheets['Counts'].cell(row = counts_max_row + 2, column = 5).value = '= SUM(E1:E' + str(counts_max_row) + ')'
    writer.sheets['Counts'].cell(row = counts_max_row + 2, column = 6).value = '= SUM(F1:F' + str(counts_max_row) + ')'
    writer.sheets['Counts'].cell(row = counts_max_row + 2, column = 7).value = '= SUM(G1:G' + str(counts_max_row) + ')'
    writer.sheets['Counts'].cell(row = counts_max_row + 2, column = 8).value = '= SUM(H1:H' + str(counts_max_row) + ')'
    writer.sheets['Counts'].cell(row = counts_max_row + 2, column = 9).value = '= SUM(I1:I' + str(counts_max_row) + ')'
    writer.sheets['Counts'].cell(row = counts_max_row + 2, column = 10).value = '= SUM(J1:J' + str(counts_max_row) + ')'
    writer.sheets['Counts'].cell(row = counts_max_row + 2, column = 11).value = '= SUM(K1:K' + str(counts_max_row) + ')'
    writer.sheets['Counts'].cell(row = counts_max_row + 2, column = 12).value =  '= SUM(L1:L' + str(counts_max_row) + ')'
    writer.sheets['Counts'].cell(row = counts_max_row + 2, column = 11).value = '= SUM(M1:M' + str(counts_max_row) + ')'
    writer.sheets['Counts'].cell(row = counts_max_row + 2, column = 12).value =  '= SUM(N1:N' + str(counts_max_row) + ')'

    #
    #
    first_sheet = workbook.get_sheet_by_name('Sheet')



    workbook.remove_sheet(first_sheet)
    writer.save()
    workbook.save(filename)

    # print(proc_dept_df_list)
    # print(proc_dept_df_list[d])

    d += 1
    # print("All Ebooks: " + str(all_ebooks))
    # print("Total ebook count: " + str(ebook_counter) + "\n")
    #sru_url_prefix = "https://tufts.alma.exlibrisgroup.com/view/sru/01TUN_INST?version=1.2&operation=searchRetrieve&recordSchema=marcxml&query=alma.title="


# x = 0
# print("Starting reading through RESERVES file\n")
# # for event, elem in et.iterparse(covid_filename):
# while x < len(reserves_df.index):
#     if x > 1:
#         break
#
#     title = reserves_df.iloc[x]['title_only']
#
#     title = re.sub(r'\s', '%20', title)
#     title = '%22' + title + '%22'
#     record = requests.get(sru_url_prefix + title)
#     print(title)
#     print(record.text)
#     x += 1

    # tree = et.ElementTree(et.fromstring(covid_xml))
    # root = tree.get_root()
    #
    # for record in root.findall('record'):

    #bib_record = pym.parse_xml_to_array(io.StringIO(elem))










# reserves_df.to_excel(oDir + '/Titles from Reserves List.xlsx', index=False)
# electronic_df.to_excel(oDir + 'Titles from Electronic List.xlsx', index=False)
# covid_df.to_excel(oDir + '/Parsed COVID Titles - Sample.xlsx', index=False)
#
#
end = datetime.datetime.now() - start
# test_file.close()

print("Execution time for creating COVID dataframe: " + str(end) + "\n")

#     x += 1
#

#
#
#
# for bib_record in bib_records:
#     if x > 5:
#         break
#     title = bib_record.get_fields('245')
#     print("Title: "  + str(title) +"\n")
#
#     x += 1
#
#
#
#
# # reserves_df['Title (Complete)'] = reserves_df['Title (Complete)'].apply(lambda x: x.lower())
# # electronic_df['Title (Complete)'] = electronic_df['Title (Complete)'].apply(lambda x: x.lower())
#
# print(reserves_df['Title (Complete)'])
# print(electronic_df['Title (Complete)'])
# print(covid_df['Title'])
# # #
# print(covid_df['title'])

#covid_sample = covid_df.head()

#covid_sample.to_excel(oDir + '/test.xlsx', index=False)
#
# proquest_df.columns = ['MMS Id', 'TitleProQuest', 'Vendor Name']
#
# overlap_df.columns = ['MMS Id', 'TitleNonProQuest', 'Vendor Name']


#### KEEP - MAYBE
# start2 = datetime.datetime.now()
# master = pd.merge(reserves_df, electronic_df, left_on=['Title (Complete)'], right_on=['Title (Complete)'], how='outer', indicator=True)
# end2 = datetime.datetime.now() - start2
#
#
# print("Execution time for merging with electronic: " + str(end2) + "\n")
# #
# electronic_overlap_df = master[master['_merge'] == 'both']
#
# start3 = datetime.datetime.now()
# master_2 = pd.merge(reserves_df, covid_df, left_on=['Title (Complete)'], right_on=['Title'], how='outer', indicator=True)
#
# end3 = datetime.datetime.now() - start3
# print("Execution time for merging with COVID: " + str(end3) + "\n")
# covid_overlap_df = master[master['_merge'] == 'both']
# #
# start4 = datetime.datetime.now()
# final_df = pd.merge(electronic_overlap_df, covid_overlap_df, on=['Title (Complete)'], how='outer')
# end4 = datetime.datetime.now() - start4
# print("Execution time for merging merged files: " + str(end4) + "\n")
#
# final_df.to_excel(oDir + '/Overlap Analsysis - Reserves.xlsx', index=False)


#### END KEEP - MAYBE
